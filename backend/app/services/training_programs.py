from io import BytesIO
from collections import defaultdict
import uuid

from fastapi import HTTPException
from fastapi import UploadFile
from openpyxl import load_workbook
from sqlmodel import Session, select, func, or_, desc
from starlette import status

from app.models.models import (
    Departments,
    Majors,
    Specializations,
    Subjects,
    TrainingProgram,
    TrainingProgramSubject,
)
from app.models.schemas.training_program.training_program_create_schemas import (
    TrainingProgramDepartmentInfo,
    TrainingProgramCreateWithSubjects,
    TrainingProgramDeleteResponse,
    TrainingProgramDropDownResponse,
    TrainingProgramListResponse,
    TrainingProgramPublic,
    TrainingProgramMajorInfo,
    TrainingProgramQueryParams,
    TrainingProgramSpecializationInfo,
    TrainingProgramSubjectDetailPublic,
    TrainingProgramSubjectPublic,
    TrainingProgramUpdateResponse,
    TrainingProgramUpdateWithSubjects,
    TrainingProgramWithSubjectsPublic,
)
from app.models.schemas.training_program.training_program_file_schemas import (
    TrainingProgramFileData,
    TrainingProgramFileDataResponse,
    TrainingProgramFileInfo,
    TrainingProgramFileInvalidSubject,
    TrainingProgramFileSubjectData,
)

class TrainingProgramServices:
    @staticmethod
    def get_dropdown(
        *, session: Session, query: TrainingProgramQueryParams
    ) -> list[TrainingProgramDropDownResponse]:
        statement = select(TrainingProgram)

        conditions = []
        if query.status:
            conditions.append(TrainingProgram.status == query.status)
        if query.specialization_id:
            conditions.append(TrainingProgram.specialization_id == query.specialization_id)
        if query.search:
            search_pattern = f"%{query.search}%"
            conditions.append(
                or_(
                    TrainingProgram.program_type.ilike(search_pattern),
                    TrainingProgram.training_program_name.ilike(search_pattern),
                    TrainingProgram.academic_year.ilike(search_pattern),
                )
            )

        if conditions:
            statement = statement.where(*conditions)

        statement = statement.order_by(desc(TrainingProgram.created_at))
        statement = statement.offset(query.skip).limit(query.limit)
        rows = session.exec(statement).all()

        return [
            TrainingProgramDropDownResponse(
                id=row.id,
                program_type=row.program_type,
                training_program_name=row.training_program_name,
                academic_year=row.academic_year,
            )
            for row in rows
        ]

    @staticmethod
    def get_dropdown_by_ids(
        *, session: Session, ids: list[uuid.UUID]
    ) -> list[TrainingProgramDropDownResponse]:
        if not ids:
            return []

        rows = session.exec(
            select(TrainingProgram).where(TrainingProgram.id.in_(ids))
        ).all()

        return [
            TrainingProgramDropDownResponse(
                id=row.id,
                program_type=row.program_type,
                training_program_name=row.training_program_name,
                academic_year=row.academic_year,
            )
            for row in rows
        ]

    @staticmethod
    def _resolve_specialization(
        *, session: Session, specialization_id, specialization_code: str | None
    ) -> Specializations:
        if specialization_id:
            specialization = session.get(Specializations, specialization_id)
            if specialization:
                return specialization

        if specialization_code:
            specialization = session.exec(
                select(Specializations).where(
                    Specializations.specialization_code == specialization_code
                )
            ).first()
            if specialization:
                return specialization

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Specialization id or specialization code is required.",
        )

    @staticmethod
    def get_training_programs(
        *, session: Session, query: TrainingProgramQueryParams
    ) -> TrainingProgramListResponse:
        statement = (
            select(
                TrainingProgram,
                Specializations.id,
                Specializations.specialization_code,
                Specializations.name,
                Majors.id,
                Majors.major_code,
                Majors.name,
                Departments.id,
                Departments.department_code,
                Departments.name,
            )
            .join(Specializations, Specializations.id == TrainingProgram.specialization_id)
            .join(Majors, Majors.id == Specializations.major_id)
            .join(Departments, Departments.id == Majors.department_id)
        )

        conditions = []
        if query.status:
            conditions.append(TrainingProgram.status == query.status)
        if query.specialization_id:
            conditions.append(TrainingProgram.specialization_id == query.specialization_id)
        if query.search:
            conditions.append(
                or_(
                    TrainingProgram.program_type.ilike(f"%{query.search}%"),
                    TrainingProgram.training_program_name.ilike(f"%{query.search}%"),
                    TrainingProgram.academic_year.ilike(f"%{query.search}%"),
                    Specializations.name.ilike(f"%{query.search}%"),
                )
            )

        if conditions:
            statement = statement.where(*conditions)

        total = session.exec(
            select(func.count()).select_from(statement.subquery())
        ).one()

        statement = statement.order_by(desc(TrainingProgram.created_at))
        statement = statement.offset(query.skip).limit(query.limit)
        rows = session.exec(statement).all()
        training_program_ids = [row[0].id for row in rows]
        subject_map: dict = defaultdict(list)

        if training_program_ids:
            subject_rows = session.exec(
                select(
                    TrainingProgramSubject,
                    Subjects.subject_code,
                    Subjects.name,
                    Subjects.credit,
                )
                .join(Subjects, Subjects.id == TrainingProgramSubject.subject_id)
                .where(TrainingProgramSubject.training_program_id.in_(training_program_ids))
                .order_by(
                    TrainingProgramSubject.training_program_id,
                    TrainingProgramSubject.term,
                )
            ).all()

            for subject_row in subject_rows:
                training_program_subject = subject_row[0]
                subject_map[training_program_subject.training_program_id].append(
                    TrainingProgramSubjectDetailPublic(
                        id=training_program_subject.id,
                        training_program_id=training_program_subject.training_program_id,
                        subject_id=training_program_subject.subject_id,
                        subject_code=subject_row[1],
                        subject_name=subject_row[2],
                        credit=subject_row[3],
                        term=training_program_subject.term,
                        status=training_program_subject.status,
                    )
                )

        data = []
        for row in rows:
            program = row[0]
            specialization_info = TrainingProgramSpecializationInfo(
                id=row[1],
                specialization_code=row[2],
                specialization_name=row[3],
            )
            major_info = TrainingProgramMajorInfo(
                id=row[4],
                major_code=row[5],
                major_name=row[6],
            )
            department_info = TrainingProgramDepartmentInfo(
                id=row[7],
                department_code=row[8],
                department_name=row[9],
            )
            data.append(
                TrainingProgramPublic(
                    id=program.id,
                    program_type=program.program_type,
                    training_program_name=program.training_program_name,
                    academic_year=program.academic_year,
                    specialization_id=program.specialization_id,
                    specialization_infor=specialization_info,
                    major_infor=major_info,
                    department_info=department_info,
                    status=program.status,
                    subjects=subject_map.get(program.id, []),
                )
            )

        return TrainingProgramListResponse(total=total, data=data)

    @staticmethod
    async def upload_file_training_program(
        *, session: Session, file: UploadFile
    ) -> TrainingProgramFileDataResponse:
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel sheet is empty.",
            )

        meta_row_indexes = {
            "program_type": 4,
            "training_program_name": 5,
            "specialization_code": 6,
            "specialization_name": 7,
            "academic_year": 8,
        }
        header_row_index = 10
        table_headers = ["TT", "Mã MH", "Tên MH", "Học kỳ", "Ghi chú"]

        if len(rows) <= header_row_index:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not contain the training program table.",
            )

        def cell(row_index: int, col_index: int) -> object | None:
            if row_index >= len(rows):
                return None
            row_values = rows[row_index]
            if row_values is None or col_index >= len(row_values):
                return None
            return row_values[col_index]

        training_program = TrainingProgramFileData(
            program_type=str(cell(meta_row_indexes["program_type"], 2) or "").strip() or None,
            training_program_name=str(cell(meta_row_indexes["training_program_name"], 2) or "").strip() or None,
            specialization_code=str(cell(meta_row_indexes["specialization_code"], 2) or "").strip() or None,
            specialization_name=str(cell(meta_row_indexes["specialization_name"], 2) or "").strip() or None,
            academic_year=str(cell(meta_row_indexes["academic_year"], 2) or "").strip() or None,
        )

        specialization = None
        if training_program.specialization_code:
            specialization = session.exec(
                select(Specializations).where(
                    Specializations.specialization_code == training_program.specialization_code
                )
            ).first()

        if not specialization:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Specialization code is required in the import file.",
            )

        training_program.specialization_code = specialization.specialization_code
        if not training_program.specialization_name:
            training_program.specialization_name = specialization.name

        parsed_subjects: list[TrainingProgramFileSubjectData] = []
        invalid_subjects: list[TrainingProgramFileInvalidSubject] = []

        for row_index, row_values in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            if row_values is None:
                continue

            def row_cell(col_index: int) -> object | None:
                if col_index >= len(row_values):
                    return None
                return row_values[col_index]

            subject_code = str(row_cell(1) or "").strip()
            subject_name = str(row_cell(2) or "").strip()
            term_raw = row_cell(3)

            if not any([subject_code, subject_name, term_raw]):
                continue

            errors: list[str] = []
            if not subject_code:
                errors.append("Subject code is required.")
            if not subject_name:
                errors.append("Subject name is required.")
            if term_raw in (None, ""):
                errors.append("Term is required.")

            term_value: int | None = None
            if term_raw not in (None, ""):
                try:
                    term_value = int(term_raw)
                    if term_value < 1:
                        raise ValueError
                except Exception:
                    errors.append("Term must be a positive integer.")

            if errors:
                invalid_subjects.append(
                    TrainingProgramFileInvalidSubject(
                        row=row_index,
                        subject_code=subject_code or None,
                        subject_name=subject_name or None,
                        term=term_value,
                        errors=errors,
                    )
                )
                continue

            parsed_subjects.append(
                TrainingProgramFileSubjectData(
                    subject_code=subject_code,
                    subject_name=subject_name,
                    term=term_value,
                )
            )

        training_program.subjects = parsed_subjects

        return TrainingProgramFileDataResponse(
            file_information=TrainingProgramFileInfo(
                file_name=filename,
                headers=table_headers,
                header_row=header_row_index + 1,
                total_rows=max(len(rows) - (header_row_index + 1), 0),
                valid_rows_count=len(parsed_subjects),
                invalid_rows_count=len(invalid_subjects),
            ),
            training_program=training_program,
            invalid_subjects=invalid_subjects,
        )

    @staticmethod
    def create_with_subjects(
        *, session: Session, payload: TrainingProgramCreateWithSubjects
    ) -> TrainingProgramWithSubjectsPublic:
        specialization = TrainingProgramServices._resolve_specialization(
            session=session,
            specialization_id=payload.specialization_id,
            specialization_code=payload.specialization_code,
        )

        existing_program = session.exec(
            select(TrainingProgram).where(
                TrainingProgram.academic_year == payload.academic_year,
                TrainingProgram.specialization_id == specialization.id,
                TrainingProgram.program_type == payload.program_type,
            )
        ).first()
        if existing_program:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Training program already exists for this specialization, academic year and program type.",
            )

        try:
            program_data = payload.model_dump(
                exclude={"subjects", "specialization_code", "specialization_name", "specialization_id"}
            )
            new_program = TrainingProgram(
                **program_data,
                specialization_id=specialization.id,
            )
            session.add(new_program)
            session.flush()

            created_subjects: list[TrainingProgramSubjectPublic] = []
            for index, subject_item in enumerate(payload.subjects, start=1):
                subject = session.exec(
                    select(Subjects).where(
                        Subjects.subject_code == subject_item.subject_code
                    )
                ).first()
                if not subject:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"Subject with code {subject_item.subject_code} does not exist.",
                    )

                existing_link = session.exec(
                    select(TrainingProgramSubject).where(
                        TrainingProgramSubject.training_program_id == new_program.id,
                        TrainingProgramSubject.subject_id == subject.id,
                    )
                ).first()
                if existing_link:
                    continue

                new_link = TrainingProgramSubject(
                    training_program_id=new_program.id,
                    subject_id=subject.id,
                    term=subject_item.term or index,
                    status=payload.status,
                )
                session.add(new_link)
                session.flush()
                created_subjects.append(
                    TrainingProgramSubjectPublic(
                        id=new_link.id,
                        training_program_id=new_link.training_program_id,
                        subject_id=new_link.subject_id,
                        term=new_link.term,
                        status=new_link.status,
                    )
                )

            session.commit()

            return TrainingProgramWithSubjectsPublic(
                id=new_program.id,
                program_type=new_program.program_type,
                training_program_name=new_program.training_program_name,
                academic_year=new_program.academic_year,
                specialization_id=new_program.specialization_id,
                status=new_program.status,
                subjects=created_subjects,
            )
        except HTTPException:
            session.rollback()
            raise
        except Exception:
            session.rollback()
            raise

    @staticmethod
    def update_with_subjects(
        *, session: Session, training_program_id: str, payload: TrainingProgramUpdateWithSubjects
    ) -> TrainingProgramUpdateResponse:
        program = session.get(TrainingProgram, training_program_id)
        if not program:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Training program not found.",
            )

        specialization = None
        if payload.specialization_id:
            specialization = session.get(Specializations, payload.specialization_id)
            if not specialization:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Specialization not found.",
                )

        try:
            update_data = payload.model_dump(exclude_unset=True, exclude={"subjects", "specialization_id"})
            for field, value in update_data.items():
                setattr(program, field, value)

            if specialization:
                program.specialization_id = specialization.id

            existing_links = session.exec(
                select(TrainingProgramSubject).where(
                    TrainingProgramSubject.training_program_id == program.id
                )
            ).all()
            existing_by_subject_id = {link.subject_id: link for link in existing_links}

            desired_subject_ids = set()
            updated_subjects: list[TrainingProgramSubjectPublic] = []

            for index, subject_item in enumerate(payload.subjects, start=1):
                subject = session.get(Subjects, subject_item.subject_id)
                if not subject:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"Subject with id {subject_item.subject_id} does not exist.",
                    )

                desired_subject_ids.add(subject.id)
                existing_link = existing_by_subject_id.get(subject.id)
                term_value = subject_item.term or index

                if existing_link:
                    existing_link.term = term_value
                    existing_link.status = payload.status if payload.status is not None else existing_link.status
                    session.add(existing_link)
                    updated_subjects.append(
                        TrainingProgramSubjectPublic(
                            id=existing_link.id,
                            training_program_id=existing_link.training_program_id,
                            subject_id=existing_link.subject_id,
                            term=existing_link.term,
                            status=existing_link.status,
                        )
                    )
                else:
                    new_link = TrainingProgramSubject(
                        training_program_id=program.id,
                        subject_id=subject.id,
                        term=term_value,
                        status=payload.status,
                    )
                    session.add(new_link)
                    session.flush()
                    updated_subjects.append(
                        TrainingProgramSubjectPublic(
                            id=new_link.id,
                            training_program_id=new_link.training_program_id,
                            subject_id=new_link.subject_id,
                            term=new_link.term,
                            status=new_link.status,
                        )
                    )

            for existing_link in existing_links:
                if existing_link.subject_id not in desired_subject_ids:
                    session.delete(existing_link)

            session.commit()
            session.refresh(program)

            specialization_row = session.exec(
                select(
                    Specializations.id,
                    Specializations.specialization_code,
                    Specializations.name,
                    Majors.id,
                    Majors.major_code,
                    Majors.name,
                    Departments.id,
                    Departments.department_code,
                    Departments.name,
                )
                .join(Majors, Majors.id == Specializations.major_id)
                .join(Departments, Departments.id == Majors.department_id)
                .where(Specializations.id == program.specialization_id)
            ).first()

            if not specialization_row:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Specialization chain not found.",
                )

            subject_rows = session.exec(
                select(
                    TrainingProgramSubject,
                    Subjects.subject_code,
                    Subjects.name,
                    Subjects.credit,
                )
                .join(Subjects, Subjects.id == TrainingProgramSubject.subject_id)
                .where(TrainingProgramSubject.training_program_id == program.id)
                .order_by(TrainingProgramSubject.term)
            ).all()

            subject_details = [
                TrainingProgramSubjectDetailPublic(
                    id=row[0].id,
                    training_program_id=row[0].training_program_id,
                    subject_id=row[0].subject_id,
                    subject_code=row[1],
                    subject_name=row[2],
                    credit=row[3],
                    term=row[0].term,
                    status=row[0].status,
                )
                for row in subject_rows
            ]

            return TrainingProgramUpdateResponse(
                id=program.id,
                program_type=program.program_type,
                training_program_name=program.training_program_name,
                academic_year=program.academic_year,
                specialization_id=program.specialization_id,
                specialization_infor=TrainingProgramSpecializationInfo(
                    id=specialization_row[0],
                    specialization_code=specialization_row[1],
                    specialization_name=specialization_row[2],
                ),
                major_infor=TrainingProgramMajorInfo(
                    id=specialization_row[3],
                    major_code=specialization_row[4],
                    major_name=specialization_row[5],
                ),
                department_info=TrainingProgramDepartmentInfo(
                    id=specialization_row[6],
                    department_code=specialization_row[7],
                    department_name=specialization_row[8],
                ),
                status=program.status,
                subjects=subject_details,
            )
        except HTTPException:
            session.rollback()
            raise
        except Exception:
            session.rollback()
            raise

    @staticmethod
    def delete(
        *, session: Session, training_program_id: str
    ) -> TrainingProgramDeleteResponse:
        program = session.get(TrainingProgram, training_program_id)
        if not program:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Training program not found.",
            )

        if program.status != "inactive":
            program.status = "inactive"
            message = "Training program set to inactive"
        else:
            message = "Training program already inactive"

        session.commit()

        return TrainingProgramDeleteResponse(id=program.id, message=message)
