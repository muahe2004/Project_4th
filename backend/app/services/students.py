import uuid
from io import BytesIO
from datetime import datetime
import logging
import re
from fastapi import HTTPException, Request, UploadFile
from sqlmodel import Session, and_, func, or_, select
from starlette import status
from app.middleware.hashing import hash_password
from typing import List
from openpyxl import load_workbook

from app.models.models import (
    Classes,
    Relatives,
    StudentClass,
    Students,
    TuitionFees,
    UserInformations,
)
from app.models.schemas.students.student_schemas import (
    StudentCreateResponse,
    StudentFileData,
    StudentFileDataResponse,
    StudentFileInfo,
    StudentFileInvalidRow,
    StudentUploadField,
    StudentPublic,
    StudentCreateWithUserInfor,
    StudentUpdate,
    StudentDeleteResponse,
    StudentsResponse,
    UserRelativeCreate,
)
from app.models.schemas.user_informations.user_information_schemas import (
    UserInformationPublic,
    UserInformationUpdate,
)
from app.models.schemas.relatives.relative_schemas import RelativePublic
from app.services.user_information import User_Information_Services
from app.services.relatives import RelativeServices
from app.services.classes import ClassServices
from app.services.common import parse_excel_datetime, to_clean_text
from app.enums.status import StatusEnum
from app.enums.class_type import ClassTypeEnum

logger = logging.getLogger(__name__)

def _has_relative_payload_value(relative: UserRelativeCreate) -> bool:
    # `relatives.name` is required in DB, so only keep records with a non-empty name.
    return bool(relative.name and str(relative.name).strip())


def _sanitize_user_information_payload(payload: dict) -> dict:
    ignored_fields = {"student_id", "teacher_id", "created_at", "updated_at", "id"}
    return {key: value for key, value in payload.items() if key not in ignored_fields}


def _normalize_gender(raw_value: object) -> str | None:
    text_value = to_clean_text(raw_value)
    if not text_value:
        return None

    normalized = text_value.strip().lower()
    if normalized == "nam":
        return "1"
    if normalized in {"nữ", "nu"}:
        return "2"
    return "3"


class StudentServices:
    @staticmethod
    def get_student_term(student_course: str) -> int | None:
        """Estimate the current term from the student's course start year and current year."""
        if not student_course:
            logger.info("Student term debug: student_course is empty -> term=None")
            return None

        match = re.search(r"(\d{4})\s*-\s*(\d{4})", student_course)
        if not match:
            logger.info(
                "Student term debug: invalid student_course=%s -> term=None",
                student_course,
            )
            return None

        start_year = int(match.group(1))
        end_year = int(match.group(2))
        current_year = datetime.now().year

        term = (current_year - start_year) * 2
        if term < 1:
            term = 1

        logger.info(
            "Student term debug: student_course=%s, start_year=%s, end_year=%s, current_year=%s, term=%s",
            student_course,
            start_year,
            end_year,
            current_year,
            term,
        )
        return term

    @staticmethod
    def _deactivate_other_primary_class_links(
        *, session: Session, student_id: uuid.UUID, keep_id: uuid.UUID | None
    ) -> None:
        conditions = [
            StudentClass.student_id == student_id,
            or_(
                StudentClass.class_type == ClassTypeEnum.PRIMARY,
                StudentClass.class_type == "TEACHER",  # dữ liệu cũ
            ),
            or_(
                StudentClass.status == StatusEnum.ACTIVE,
                StudentClass.status.is_(None),
            ),
        ]
        if keep_id is not None:
            conditions.append(StudentClass.id != keep_id)

        rows = session.exec(select(StudentClass).where(and_(*conditions))).all()
        for row in rows:
            row.status = StatusEnum.INACTIVE
            row.updated_at = datetime.now()

    @staticmethod
    def _get_primary_class_map(
        *, session: Session, student_ids: List[uuid.UUID]
    ) -> dict[uuid.UUID, uuid.UUID]:
        if not student_ids:
            return {}

        student_class_rows = session.exec(
            select(StudentClass)
            .where(StudentClass.student_id.in_(student_ids))
            .order_by(StudentClass.updated_at.desc(), StudentClass.created_at.desc())
        ).all()

        fallback_map: dict[uuid.UUID, uuid.UUID] = {}
        active_map: dict[uuid.UUID, uuid.UUID] = {}

        for row in student_class_rows:
            if row.student_id is None:
                continue

            if row.student_id not in fallback_map:
                fallback_map[row.student_id] = row.class_id

            if (
                row.student_id not in active_map
                and row.status in (None, StatusEnum.ACTIVE)
            ):
                active_map[row.student_id] = row.class_id

        result = fallback_map.copy()
        result.update(active_map)
        return result

    @staticmethod
    def _get_primary_class_id(
        *, session: Session, student_id: uuid.UUID
    ) -> uuid.UUID | None:
        return StudentServices._get_primary_class_map(
            session=session, student_ids=[student_id]
        ).get(student_id)

    @staticmethod
    def _upsert_student_class(
        *,
        session: Session,
        student_id: uuid.UUID,
        class_id: uuid.UUID,
        set_primary_class_type: bool = False,
        replace_primary_class: bool = False,
    ) -> None:
        if replace_primary_class:
            primary_link = session.exec(
                select(StudentClass)
                .where(
                    StudentClass.student_id == student_id,
                    or_(
                        StudentClass.class_type == ClassTypeEnum.PRIMARY,
                        StudentClass.class_type == "TEACHER",  # dữ liệu cũ
                    ),
                )
                .order_by(StudentClass.updated_at.desc(), StudentClass.created_at.desc())
            ).first()
            if primary_link:
                primary_link.class_id = class_id
                primary_link.status = StatusEnum.ACTIVE
                if set_primary_class_type:
                    primary_link.class_type = ClassTypeEnum.PRIMARY
                    StudentServices._deactivate_other_primary_class_links(
                        session=session,
                        student_id=student_id,
                        keep_id=primary_link.id,
                    )
                primary_link.updated_at = datetime.now()
                return

        existing_link = session.exec(
            select(StudentClass).where(
                StudentClass.student_id == student_id,
                StudentClass.class_id == class_id,
            )
        ).first()

        if existing_link:
            existing_link.status = StatusEnum.ACTIVE
            if set_primary_class_type and not existing_link.class_type:
                existing_link.class_type = ClassTypeEnum.PRIMARY
            if set_primary_class_type and replace_primary_class:
                StudentServices._deactivate_other_primary_class_links(
                    session=session,
                    student_id=student_id,
                    keep_id=existing_link.id,
                )
            existing_link.updated_at = datetime.now()
            return

        class_type = ClassTypeEnum.PRIMARY if set_primary_class_type else None
        if set_primary_class_type and replace_primary_class:
            StudentServices._deactivate_other_primary_class_links(
                session=session,
                student_id=student_id,
                keep_id=None,
            )
        session.add(
            StudentClass(
                student_id=student_id,
                class_id=class_id,
                status=StatusEnum.ACTIVE,
                class_type=class_type,
            )
        )

    @staticmethod
    def get_all(*, session: Session, query) -> tuple[list[StudentsResponse], int]:
        statement = select(Students)

        conditions = []
        if query.status:
            conditions.append(Students.status == query.status)

        if query.class_id:
            student_ids_by_class = session.exec(
                select(StudentClass.student_id).where(
                    StudentClass.class_id == query.class_id,
                    or_(
                        StudentClass.status == StatusEnum.ACTIVE,
                        StudentClass.status.is_(None),
                    ),
                )
            ).all()
            filtered_student_ids = [
                student_id
                for student_id in student_ids_by_class
                if student_id is not None
            ]
            if not filtered_student_ids:
                return [], 0
            conditions.append(Students.id.in_(filtered_student_ids))

        if query.search:
            conditions.append(
                or_(
                    Students.student_code.ilike(f"%{query.search}%"),
                    Students.name.ilike(f"%{query.search}%"),
                )
            )

        if conditions:
            statement = statement.where(*conditions)

        count_stmt = select(func.count()).select_from(Students)
        if conditions:
            count_stmt = count_stmt.where(*conditions)
        total = session.exec(count_stmt).one()

        statement = (
            statement.order_by(Students.created_at.desc())
            .offset(query.skip)
            .limit(query.limit)
        )

        students_page = session.exec(statement).all()
        student_ids = [student.id for student in students_page]

        user_infos = {}
        if student_ids:
            infos = session.exec(
                select(UserInformations).where(
                    UserInformations.student_id.in_(student_ids)
                )
            ).all()
            user_infos = {info.student_id: info for info in infos}

        relatives_map = {}
        if student_ids:
            relatives = session.exec(
                select(Relatives).where(Relatives.student_id.in_(student_ids))
            ).all()
            for relative in relatives:
                relatives_map.setdefault(relative.student_id, []).append(relative)

        student_class_map = StudentServices._get_primary_class_map(
            session=session,
            student_ids=student_ids,
        )
        class_ids = [
            class_id for class_id in student_class_map.values() if class_id is not None
        ]

        class_info = ClassServices.get_dropdown_by_ids(
            session=session,
            ids=class_ids,
            request=None,
        )
        class_map = {
            str(c.id): {
                "class_code": c.class_code,
                "class_name": c.class_name,
            }
            for c in class_info
        }

        students: list[StudentsResponse] = []
        for student in students_page:
            user_info = user_infos.get(student.id)
            student_relatives = relatives_map.get(student.id, [])
            class_id = student_class_map.get(student.id)
            class_data = class_map.get(str(class_id), {})

            students.append(
                StudentsResponse(
                    id=student.id,
                    student_code=student.student_code,
                    name=student.name,
                    date_of_birth=student.date_of_birth,
                    gender=student.gender,
                    email=student.email,
                    phone=student.phone,
                    address=student.address,
                    class_id=class_id,
                    training_program=student.training_program,
                    course=student.course,
                    status=student.status,
                    created_at=student.created_at,
                    updated_at=student.updated_at,
                    class_code=class_data.get("class_code"),
                    class_name=class_data.get("class_name"),
                    student_information=(
                        UserInformationPublic.model_validate(user_info)
                        if user_info is not None
                        else None
                    ),
                    student_relative=student_relatives,
                )
            )

        return students, total

    @staticmethod
    def create(
        *, session: Session, student: StudentCreateWithUserInfor
    ) -> StudentCreateResponse:

        if session.exec(
            select(Students).where(Students.student_code == student.student_code)
        ).first():
            raise HTTPException(400, "Student already exists")

        student_data = student.model_dump(
            exclude={"student_information", "student_relatives", "class_id"}
        )
        student_data["password"] = hash_password(student_data["password"])

        new_student = Students(**student_data)
        session.add(new_student)
        session.flush()

        user_info_payload = _sanitize_user_information_payload(
            student.student_information.model_dump(exclude_none=True)
        )
        user_info = UserInformations(**user_info_payload, student_id=new_student.id)

        session.add(user_info)

        relatives = []
        for relative in student.student_relatives:
            relative_data = relative.model_dump(exclude_none=True)
            if not relative_data:
                continue
            relative_record = Relatives(
                **relative_data,
                student_id=new_student.id,
            )
            session.add(relative_record)
            relatives.append(relative_record)

        if student.class_id is not None:
            StudentServices._upsert_student_class(
                session=session,
                student_id=new_student.id,
                class_id=student.class_id,
                set_primary_class_type=True,
            )

        session.commit()
        session.refresh(new_student)
        session.refresh(user_info)

        student_relatives_response = (
            [RelativePublic.model_validate(rel) for rel in relatives]
            if relatives
            else None
        )
        return StudentCreateResponse(
            **new_student.dict(),
            class_id=student.class_id,
            student_information=UserInformationPublic.model_validate(user_info),
            student_relative=student_relatives_response,
        )

    @staticmethod
    def get_by_id(
        *, session: Session, student_id: uuid.UUID, request: Request
    ) -> StudentPublic:
        student = session.get(Students, student_id)
        if not student:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Student does not exist"
            )
        class_id = StudentServices._get_primary_class_id(
            session=session,
            student_id=student.id,
        )
        payload = student.model_dump()
        payload["class_id"] = class_id
        return StudentPublic.model_validate(payload)

    @staticmethod
    def create_list_student() -> StudentPublic:
        return "Create list student"

    @staticmethod
    def update(
        *,
        session: Session,
        student_id: uuid.UUID,
        student_data: StudentUpdate,
    ) -> StudentPublic:
        student = session.get(Students, student_id)
        if not student:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
            )

        update_data = student_data.model_dump(
            exclude_unset=True,
            exclude={"student_information", "student_relatives", "class_id"},
        )
        for field, value in update_data.items():
            setattr(student, field, value)

        if student_data.class_id is not None:
            StudentServices._upsert_student_class(
                session=session,
                student_id=student.id,
                class_id=student_data.class_id,
                set_primary_class_type=True,
                replace_primary_class=True,
            )

        if student_data.student_information:
            info_payload = _sanitize_user_information_payload(
                student_data.student_information.model_dump(exclude_none=True)
            )
            if info_payload:
                user_info = session.exec(
                    select(UserInformations).where(
                        UserInformations.student_id == student.id
                    )
                ).one_or_none()
                user_info_update = UserInformationUpdate(**info_payload)
                if user_info:
                    User_Information_Services.update(
                        session=session,
                        user_information_id=user_info.id,
                        user_information_data=user_info_update,
                        commit=False,
                    )
                else:
                    session.add(UserInformations(**info_payload, student_id=student.id))

        if student_data.student_relatives is not None:
            filtered_relatives = [
                rel
                for rel in student_data.student_relatives
                if _has_relative_payload_value(rel)
            ]
            RelativeServices.replace_for_student(
                session=session,
                student_id=student.id,
                relatives=filtered_relatives,
                commit=False,
            )

        session.commit()
        class_id = StudentServices._get_primary_class_id(
            session=session,
            student_id=student.id,
        )
        payload = student.model_dump()
        payload["class_id"] = class_id
        return StudentPublic.model_validate(payload)

    @staticmethod
    def delete_many(
        *, session: Session, student_ids: List[uuid.UUID]
    ) -> List[StudentDeleteResponse]:
        results: List[StudentDeleteResponse] = []

        for student_id in student_ids:
            student = session.get(Students, student_id)
            if not student:
                results.append(
                    StudentDeleteResponse(
                        id=str(student_id), message="Student not found"
                    )
                )
                continue

            check_related_entities = select(TuitionFees).where(
                TuitionFees.student_id == student.id
            )
            tuition_fees = session.exec(check_related_entities).all()
            if tuition_fees:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Student has related tuition fees and cannot be deleted.",
                )

            if student.status != StatusEnum.INACTIVE:
                student.status = StatusEnum.INACTIVE
                message = "Student set to inactive"
            else:
                message = "Student already inactive"

            session.commit()
            results.append(StudentDeleteResponse(id=str(student_id), message=message))

        return results

    @staticmethod
    async def upload_file_student(
        *,
        session: Session,
        file: UploadFile,
    ) -> StudentFileDataResponse:
        # validate file extension
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        # read file content
        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        # open workbook and get active sheet
        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active

        # read all rows from sheet
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel sheet is empty.",
            )

        # mapping header to field
        expected_header_labels = {
            StudentUploadField.CODE: "Mã SV",
            StudentUploadField.NAME: "Họ và tên",
            StudentUploadField.GENDER: "GT",
            StudentUploadField.CLASS_CODE: "Lớp",
            StudentUploadField.DATE_OF_BIRTH: "Ngày sinh",
            StudentUploadField.ADDRESS: "Nơi sinh",
            StudentUploadField.PHONE: "Điện thoại",
            StudentUploadField.EMAIL: "Email",
        }
        header_row_index = 6
        if len(rows) < header_row_index:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"File does not contain header row {header_row_index}. "
                    "Please use the correct template."
                ),
            )

        # student_code -> column 1, name -> column 2, class_code -> column 4,...
        fixed_column_indexes = {
            StudentUploadField.CODE: 1,
            StudentUploadField.NAME: 2,
            StudentUploadField.GENDER: 3,
            StudentUploadField.CLASS_CODE: 4,
            StudentUploadField.DATE_OF_BIRTH: 5,
            StudentUploadField.ADDRESS: 6,
            StudentUploadField.PHONE: 7,
            StudentUploadField.EMAIL: 8,
        }

        parsed_rows: list[StudentFileData] = []
        invalid_rows: list[StudentFileInvalidRow] = []

        # parse each data row below header row
        for row_index, row_values in enumerate(rows[header_row_index:], start=header_row_index + 1):
            if row_values is None:
                continue

            # get value by fixed template column index
            def cell(field_name: StudentUploadField) -> object | None:
                col_index = fixed_column_indexes[field_name]
                if col_index >= len(row_values):
                    return None
                return row_values[col_index]

            student_code = to_clean_text(cell(StudentUploadField.CODE))
            student_name = to_clean_text(cell(StudentUploadField.NAME))
            gender = _normalize_gender(cell(StudentUploadField.GENDER))
            date_of_birth_raw = cell(StudentUploadField.DATE_OF_BIRTH)
            email = to_clean_text(cell(StudentUploadField.EMAIL))
            phone = to_clean_text(cell(StudentUploadField.PHONE))
            address = to_clean_text(cell(StudentUploadField.ADDRESS))
            class_code = to_clean_text(cell(StudentUploadField.CLASS_CODE))
            class_id = None
            class_name = None
            if class_code:
                class_record = session.exec(
                    select(Classes).where(Classes.class_code == class_code)
                ).first()
                if class_record:
                    class_id = class_record.id
                    class_name = class_record.class_name

            # skip rows that are fully empty in target columns
            if all(
                value in (None, "")
                for value in (
                    student_code,
                    student_name,
                    date_of_birth_raw,
                    email,
                    phone,
                    address,
                    class_code,
                )
            ):
                continue

            # parse and validate date field
            try:
                date_of_birth = parse_excel_datetime(date_of_birth_raw)
            except ValueError as exc:
                invalid_rows.append(
                    StudentFileInvalidRow(
                        row=row_index,
                        student_code=student_code,
                        name=student_name,
                        gender=gender,
                        date_of_birth=None,
                        email=email,
                        phone=phone,
                        address=address,
                        class_id=class_id,
                        class_code=class_code,
                        class_name=class_name,
                        errors=[str(exc)],
                    )
                )
                continue

            # validate required fields
            row_errors: list[str] = []
            if not student_code:
                row_errors.append("Student Code is required.")
            if not student_name:
                row_errors.append("Student Name is required.")
            if not gender:
                row_errors.append("Gender is required.")
            if not email:
                row_errors.append("Email is required.")
            if not class_code:
                row_errors.append("Class Code is required.")

            if row_errors:
                invalid_rows.append(
                    StudentFileInvalidRow(
                        row=row_index,
                        student_code=student_code,
                        name=student_name,
                        gender=gender,
                        date_of_birth=date_of_birth,
                        email=email,
                        phone=phone,
                        address=address,
                        class_id=class_id,
                        class_code=class_code,
                        class_name=class_name,
                        errors=row_errors,
                    )
                )
                continue

            parsed_rows.append(
                StudentFileData(
                    student_code=student_code,
                    name=student_name,
                    gender=gender,
                    date_of_birth=date_of_birth,
                    email=email,
                    phone=phone,
                    address=address,
                    class_id=class_id,
                    class_code=class_code,
                    class_name=class_name,
                )
            )

        return StudentFileDataResponse(
            file_information=StudentFileInfo(
                file_name=filename,
                headers=list(expected_header_labels.values()),
                header_row=header_row_index,
                total_rows=max(len(rows) - header_row_index, 0),
                valid_rows_count=len(parsed_rows),
                invalid_rows_count=len(invalid_rows),
            ),
            students=parsed_rows,
            invalid_students=invalid_rows,
        )

    @staticmethod
    def import_list_student(
        *,
        session: Session,
        list_student: List[StudentFileData],
    ) -> List[StudentFileData]:
        imported_students: list[StudentFileData] = []

        for student_item in list_student:
            student_code = (student_item.student_code or "").strip()
            student_name = (student_item.name or "").strip()
            email = (student_item.email or "").strip()
            gender = (student_item.gender or "").strip()

            if not student_code or not student_name or not email or not gender:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="student_code, name, gender, email are required.",
                )
            if gender not in {"1", "2", "3"}:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="gender must be one of: 1, 2, 3.",
                )

            existed_student = session.exec(
                select(Students).where(
                    or_(
                        Students.student_code == student_code,
                        Students.email == email,
                    )
                )
            ).first()
            if existed_student:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Student already exists: {student_code} / {email}",
                )

            new_student = Students(
                student_code=student_code,
                name=student_name,
                date_of_birth=student_item.date_of_birth,
                gender=gender,
                email=email,
                phone=student_item.phone,
                address=student_item.address,
                training_program=None,
                course=None,
                status=StatusEnum.ACTIVE,
                password=hash_password(student_code),
            )
            session.add(new_student)
            session.flush()

            # resolve class_id by priority:
            # 1) use provided class_id
            # 2) fallback query by class_code
            resolved_class_id = student_item.class_id
            resolved_class_code = student_item.class_code
            resolved_class_name = student_item.class_name

            if resolved_class_id is None and resolved_class_code:
                class_record = session.exec(
                    select(Classes).where(Classes.class_code == resolved_class_code)
                ).first()
                if class_record:
                    resolved_class_id = class_record.id
                    resolved_class_name = class_record.class_name

            # only insert student_class when class_id is resolved
            if resolved_class_id is not None:
                session.add(
                    StudentClass(
                        student_id=new_student.id,
                        class_id=resolved_class_id,
                        status=StatusEnum.ACTIVE,
                        class_type=ClassTypeEnum.PRIMARY,
                    )
                )

            imported_students.append(
                StudentFileData(
                    student_code=new_student.student_code,
                    name=new_student.name,
                    gender=new_student.gender,
                    date_of_birth=new_student.date_of_birth,
                    email=new_student.email,
                    phone=new_student.phone,
                    address=new_student.address,
                    class_id=resolved_class_id,
                    class_code=resolved_class_code,
                    class_name=resolved_class_name,
                )
            )

        session.commit()
        return imported_students
