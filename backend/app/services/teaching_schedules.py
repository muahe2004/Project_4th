from io import BytesIO
from datetime import datetime, timedelta
import uuid
from app.enums.status import StatusEnum
from fastapi import HTTPException, Request, UploadFile
from openpyxl import load_workbook
from sqlmodel import Session, and_, desc, func, or_, select
from sqlalchemy import String, cast
from starlette import status
from typing import List, Tuple

from app.models.models import (
    Classes,
    LearningSchedules,
    Rooms,
    Subjects,
    Teachers,
    TeachingSchedules,
    StudentClass,
)
from app.models.schemas.learning_schedules.learning_schedule_schemas import (
    LearningSchedulePublic,
)
from app.models.schemas.common.query import DateRange
from app.models.schemas.teaching_schedules.teaching_schedule_schemas import (
    ImportTeachingCalenderInput,
    ImportTeachingCalenderImportedItem,
    ImportTeachingCalenderResponse,
    TeachingScheduleClassInfo,
    TeachingScheduleRoomInfo,
    TeachingScheduleResponse,
    TeachingScheduleSearchParams,
    TeachingSchedulPublic,
    TeachingScheduleSubjectInfo,
    TeachingScheduleTeacherInfo,
    TeachingScheduleCreate,
    TeachingScheduleUpdate,
    TeachingScheduleDeleteResponse,
    TeachingScheduleWithLearningSchedulePublic,
    UploadTeachingCalenderFileInfo,
    UploadTeachingCalenderInvalidRow,
    UploadTeachingCalenderItem,
    UploadTeachingCalenderResponse,
)
from app.services.learning_schedules import LearningScheduleServices
from app.services.common import build_date_conditions, to_clean_text


class TeachingScheduleServices:
    @staticmethod
    def _validate_weekday_number(weekday_number: int) -> None:
        if weekday_number not in {2, 3, 4, 5, 6, 7, 8}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="weekday must be one of: 2, 3, 4, 5, 6, 7, 8.",
            )

    @staticmethod
    def _weekday_number(date_value: datetime) -> int:
        weekday_map = {
            0: 2,
            1: 3,
            2: 4,
            3: 5,
            4: 6,
            5: 7,
            6: 8,
        }
        return weekday_map[date_value.weekday()]

    @staticmethod
    def _parse_lesson_periods(lesson_periods: str) -> list[tuple[int, int]]:
        blocks: list[tuple[int, int]] = []
        block_start: int | None = None

        for index, char in enumerate(lesson_periods, start=1):
            if char != "-" and block_start is None:
                block_start = index
            elif char == "-" and block_start is not None:
                blocks.append((block_start, index - 1))
                block_start = None

        if block_start is not None:
            blocks.append((block_start, len(lesson_periods)))

        if not blocks:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="lesson_periods must contain at least one study period.",
            )
        return blocks

    @staticmethod
    def _parse_study_week_indexes(study_weeks: str) -> list[int]:
        indexes = [index for index, char in enumerate(study_weeks, start=1) if char != "-"]
        if not indexes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="study_weeks must contain at least one study week.",
            )
        return indexes

    @staticmethod
    def _build_duplicate_schedule_message(
        *,
        class_code: str,
        subject_code: str,
        subject_name: str,
        date_value: datetime,
        start_period: int,
        end_period: int,
    ) -> str:
        return (
            f"Lớp {class_code} đã có lịch học môn {subject_code} - {subject_name} "
            f"vào ngày {date_value.strftime('%d/%m/%Y')} từ tiết {start_period} đến tiết {end_period}"
        )

    @staticmethod
    def _resolve_date_for_week(
        *,
        period_start_date: datetime,
        week_index: int,
        weekday_number: int,
    ) -> datetime:
        start_weekday_number = TeachingScheduleServices._weekday_number(period_start_date)
        weekday_offset = (weekday_number - start_weekday_number) % 7
        return period_start_date + timedelta(days=((week_index - 1) * 7) + weekday_offset)

    @staticmethod
    def get_all(
        *, session: Session, query: TeachingScheduleSearchParams, date_range: DateRange
    ) -> Tuple[List[TeachingScheduleResponse], int]:
        statement = (
            select(
                TeachingSchedules,
                LearningSchedules,
                Classes.class_name.label("class_name"),
                Classes.class_code.label("class_code"),
                Classes.id.label("class_id"),
                Teachers.name.label("teacher_name"),
                Teachers.email.label("teacher_email"),
                Teachers.phone.label("teacher_phone"),
                Teachers.id.label("teacher_id"),
                Rooms.id.label("room_id"),
                Rooms.room_number.label("room_number"),
                Subjects.id.label("subject_id"),
                Subjects.name.label("subject_name"),
                Subjects.subject_code.label("subject_code"),
            )
            .join(
                LearningSchedules,
                LearningSchedules.id == TeachingSchedules.learning_schedule_id,
            )
            .join(Classes, Classes.id == LearningSchedules.class_id)
            .join(Subjects, Subjects.id == LearningSchedules.subject_id)
            .outerjoin(Teachers, Teachers.id == TeachingSchedules.teacher_id)
            .outerjoin(Rooms, Rooms.id == LearningSchedules.room_id)
        )
        student_class_subquery = None
        if query.student_id:
            student_class_subquery = (
                select(StudentClass.class_id)
                .where(
                    StudentClass.student_id == query.student_id,
                    or_(
                        StudentClass.status == StatusEnum.ACTIVE,
                        StudentClass.status.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )
            statement = statement.join(
                student_class_subquery,
                student_class_subquery.c.class_id == LearningSchedules.class_id,
            )

        conditions = []
        if query.status:
            conditions.append(TeachingSchedules.status == query.status)

        if query.class_id:
            conditions.append(LearningSchedules.class_id == query.class_id)

        if query.teacher_id:
            conditions.append(TeachingSchedules.teacher_id == query.teacher_id)

        conditions.extend(build_date_conditions(date_range))

        if query.search:
            search_pattern = f"%{query.search}%"
            conditions.append(
                or_(
                    Classes.class_code.ilike(search_pattern),
                    Classes.class_name.ilike(search_pattern),
                    Teachers.name.ilike(search_pattern),
                    Teachers.email.ilike(search_pattern),
                    Teachers.phone.ilike(search_pattern),
                    Subjects.subject_code.ilike(search_pattern),
                    Subjects.name.ilike(search_pattern),
                    cast(Rooms.room_number, String).ilike(search_pattern),
                )
            )

        if conditions:
            statement = statement.where(and_(*conditions))

        count_stmt = (
            select(func.count())
            .select_from(TeachingSchedules)
            .join(
                LearningSchedules,
                LearningSchedules.id == TeachingSchedules.learning_schedule_id,
            )
            .join(Classes, Classes.id == LearningSchedules.class_id)
            .join(Subjects, Subjects.id == LearningSchedules.subject_id)
            .outerjoin(Teachers, Teachers.id == TeachingSchedules.teacher_id)
            .outerjoin(Rooms, Rooms.id == LearningSchedules.room_id)
        )
        if student_class_subquery is not None:
            count_stmt = count_stmt.join(
                student_class_subquery,
                student_class_subquery.c.class_id == LearningSchedules.class_id,
            )
        if conditions:
            count_stmt = count_stmt.where(and_(*conditions))
        total = session.exec(count_stmt).one()

        paged_statement = (
            statement.order_by(desc(TeachingSchedules.created_at))
            .offset(query.skip)
            .limit(query.limit)
        )

        rows = session.exec(paged_statement).all()

        schedules: List[TeachingScheduleResponse] = []
        for (
            teaching_schedule,
            learning_schedule,
            class_name,
            class_code,
            class_id,
            teacher_name,
            teacher_email,
            teacher_phone,
            teacher_id,
            room_id,
            room_number,
            subject_id,
            subject_name,
            subject_code,
        ) in rows:
            schedules.append(
                TeachingScheduleResponse(
                    id=teaching_schedule.id,
                    status=teaching_schedule.status,
                    created_at=teaching_schedule.created_at,
                    updated_at=teaching_schedule.updated_at,
                    learning_schedule=LearningSchedulePublic.model_validate(
                        learning_schedule
                    ),
                    class_info=TeachingScheduleClassInfo(
                        class_id=class_id,
                        class_name=class_name,
                        class_code=class_code,
                    ),
                    teacher=(
                        TeachingScheduleTeacherInfo(
                            teacher_id=teacher_id,
                            teacher_name=teacher_name,
                            teacher_email=teacher_email,
                            teacher_phone=teacher_phone,
                        )
                        if teacher_id is not None
                        else None
                    ),
                    room=(
                        TeachingScheduleRoomInfo(
                            room_id=room_id,
                            room_number=room_number,
                        )
                        if room_id is not None
                        else None
                    ),
                    subject=TeachingScheduleSubjectInfo(
                        subject_id=subject_id,
                        subject_name=subject_name,
                        subject_code=subject_code,
                    ),
                )
            )

        return schedules, total

    @staticmethod
    def get_by_id(
        *, session: Session, teaching_schedule_id: uuid.UUID, request: Request
    ) -> TeachingSchedulPublic:
        teaching_schedules = session.get(TeachingSchedules, teaching_schedule_id)
        if not teaching_schedules:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teaching Schedules does not exist",
            )
        return TeachingSchedulPublic.model_validate(teaching_schedules)

    @staticmethod
    def create(
        *, session: Session, teaching_schedule: TeachingScheduleCreate
    ) -> TeachingScheduleWithLearningSchedulePublic:
        if teaching_schedule.teacher_id:
            existing_teacher = session.get(Teachers, teaching_schedule.teacher_id)
            if not existing_teacher:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Teacher does not exist.",
                )

        try:
            new_learning_schedule = LearningScheduleServices.create(
                session=session,
                learning_schedules=teaching_schedule.learning_schedule,
                auto_commit=False,
            )

            new_teaching_schedule = TeachingSchedules(
                teacher_id=teaching_schedule.teacher_id,
                learning_schedule_id=new_learning_schedule.id,
                status=teaching_schedule.status,
            )
            session.add(new_teaching_schedule)
            session.commit()
            session.refresh(new_teaching_schedule)
            return TeachingScheduleWithLearningSchedulePublic(
                id=new_teaching_schedule.id,
                teacher_id=new_teaching_schedule.teacher_id,
                learning_schedule_id=new_teaching_schedule.learning_schedule_id,
                status=new_teaching_schedule.status,
                created_at=new_teaching_schedule.created_at,
                updated_at=new_teaching_schedule.updated_at,
                learning_schedule=new_learning_schedule,
            )
        except HTTPException:
            session.rollback()
            raise
        except Exception:
            session.rollback()
            raise

    @staticmethod
    def update(
        *,
        session: Session,
        teaching_schedule_id: uuid.UUID,
        teaching_schedules_data: TeachingScheduleUpdate,
    ) -> TeachingSchedulPublic:
        teaching_schedule = session.get(TeachingSchedules, teaching_schedule_id)
        if not teaching_schedule:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teaching Schedule not found",
            )

        update_data = teaching_schedules_data.model_dump(exclude_unset=True)
        learning_schedule_payload = update_data.pop("learning_schedule", None)

        if "teacher_id" in update_data:
            teacher_id = update_data["teacher_id"]
            if teacher_id is not None:
                teacher = session.get(Teachers, teacher_id)
                if not teacher:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail="Teacher does not exist.",
                    )
            teaching_schedule.teacher_id = teacher_id

        if "status" in update_data:
            teaching_schedule.status = update_data["status"]

        if learning_schedule_payload is not None:
            learning_schedule = session.get(
                LearningSchedules, teaching_schedule.learning_schedule_id
            )
            if not learning_schedule:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Learning Schedule not found",
                )

            class_id = learning_schedule_payload.get("class_id")
            if class_id is not None and not session.get(Classes, class_id):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Class does not exist.",
                )

            subject_id = learning_schedule_payload.get("subject_id")
            if subject_id is not None and not session.get(Subjects, subject_id):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Subject does not exist.",
                )

            room_id = learning_schedule_payload.get("room_id")
            if room_id is not None and not session.get(Rooms, room_id):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Room does not exist.",
                )

            start_period = learning_schedule_payload.get(
                "start_period", learning_schedule.start_period
            )
            end_period = learning_schedule_payload.get(
                "end_period", learning_schedule.end_period
            )
            if start_period is not None and end_period is not None and end_period < start_period:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="end_period must be greater than or equal to start_period.",
                )

            for field, value in learning_schedule_payload.items():
                setattr(learning_schedule, field, value)
            learning_schedule.updated_at = datetime.now()

        teaching_schedule.updated_at = datetime.now()

        session.commit()
        session.refresh(teaching_schedule)

        return TeachingSchedulPublic.model_validate(teaching_schedule)

    @staticmethod
    def delete(
        *, session: Session, teaching_schedule_id: uuid.UUID
    ) -> TeachingScheduleDeleteResponse:
        teaching_schedule = session.get(TeachingSchedules, teaching_schedule_id)
        if not teaching_schedule:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teaching Schedule not found",
            )

        learning_schedule = None
        if teaching_schedule.learning_schedule_id:
            learning_schedule = session.get(
                LearningSchedules, teaching_schedule.learning_schedule_id
            )

        if teaching_schedule.status == StatusEnum.ACTIVE:
            teaching_schedule.status = StatusEnum.INACTIVE
            if learning_schedule and learning_schedule.status == StatusEnum.ACTIVE:
                learning_schedule.status = StatusEnum.INACTIVE
            session.commit()
            return TeachingScheduleDeleteResponse(
                id=str(teaching_schedule.id),
                message="Teaching Schedule and related Learning Schedule set to inactive",
            )

        session.delete(teaching_schedule)

        if learning_schedule:
            has_other_teaching_schedule = session.exec(
                select(TeachingSchedules.id).where(
                    TeachingSchedules.learning_schedule_id == learning_schedule.id,
                    TeachingSchedules.id != teaching_schedule_id,
                )
            ).first()

            if not has_other_teaching_schedule:
                session.delete(learning_schedule)

        session.commit()

        return TeachingScheduleDeleteResponse(
            id=str(teaching_schedule.id),
            message="Teaching Schedule and related Learning Schedule deleted successfully",
        )

    @staticmethod
    async def upload_file_calender(
        *,
        session: Session,
        file: UploadFile,
    ) -> UploadTeachingCalenderResponse:
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel sheet is empty.",
            )

        expected_headers = [
            "TT",
            "Mã MH",
            "Tên MH",
            "Mã GV",
            "Tên GV",
            "Thứ",
            "Tiết Học",
            "Tuần học",
            "Phòng",
        ]
        header_row_index = 9
        if len(rows) < header_row_index:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not contain the expected header row.",
            )

        subject_codes = {
            to_clean_text(row[1])
            for row in rows[header_row_index + 1:]
            if row and len(row) > 1 and to_clean_text(row[1])
        }
        teacher_codes = {
            to_clean_text(row[3])
            for row in rows[header_row_index + 1:]
            if row and len(row) > 3 and to_clean_text(row[3])
        }
        room_numbers = {
            int(room_value)
            for row in rows[header_row_index + 1:]
            if row and len(row) > 8 and (room_value := to_clean_text(row[8])) and str(room_value).isdigit()
        }

        subject_records = session.exec(
            select(Subjects).where(Subjects.subject_code.in_(list(subject_codes)))
        ).all() if subject_codes else []
        teacher_records = session.exec(
            select(Teachers).where(Teachers.teacher_code.in_(list(teacher_codes)))
        ).all() if teacher_codes else []
        room_records = session.exec(
            select(Rooms).where(Rooms.room_number.in_(list(room_numbers)))
        ).all() if room_numbers else []

        subject_map = {subject.subject_code: subject for subject in subject_records}
        teacher_map = {teacher.teacher_code: teacher for teacher in teacher_records}
        room_map = {room.room_number: room for room in room_records}

        parsed_rows: list[UploadTeachingCalenderItem] = []
        invalid_rows: list[UploadTeachingCalenderInvalidRow] = []

        for row_index, row_values in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            if row_values is None:
                continue

            def cell(col_index: int) -> object | None:
                if col_index >= len(row_values):
                    return None
                return row_values[col_index]

            subject_code = to_clean_text(cell(1))
            subject_name = to_clean_text(cell(2))
            teacher_code = to_clean_text(cell(3))
            teacher_name = to_clean_text(cell(4))
            weekday_raw = to_clean_text(cell(5))
            lesson_periods = to_clean_text(cell(6))
            study_weeks = to_clean_text(cell(7))
            room_raw = to_clean_text(cell(8))

            if all(
                value in (None, "")
                for value in (
                    subject_code,
                    subject_name,
                    teacher_code,
                    teacher_name,
                    weekday_raw,
                    lesson_periods,
                    study_weeks,
                    room_raw,
                )
            ):
                continue

            row_errors: list[str] = []
            weekday_number: int | None = None
            room_number: int | None = None

            if not subject_code:
                row_errors.append("Subject Code is required.")
            if not teacher_code:
                row_errors.append("Teacher Code is required.")
            if not weekday_raw:
                row_errors.append("Weekday is required.")
            if not lesson_periods:
                row_errors.append("Lesson periods are required.")
            if not study_weeks:
                row_errors.append("Study weeks are required.")
            if not room_raw:
                row_errors.append("Room is required.")

            if weekday_raw:
                try:
                    weekday_number = int(weekday_raw)
                    TeachingScheduleServices._validate_weekday_number(weekday_number)
                except (ValueError, HTTPException):
                    row_errors.append("Weekday must be one of: 2, 3, 4, 5, 6, 7, 8.")

            if room_raw:
                if str(room_raw).isdigit():
                    room_number = int(room_raw)
                else:
                    row_errors.append("Room must be a number.")

            subject_record = subject_map.get(subject_code) if subject_code else None
            teacher_record = teacher_map.get(teacher_code) if teacher_code else None
            room_record = room_map.get(room_number) if room_number is not None else None

            if subject_code and subject_record is None:
                row_errors.append(f"Subject not found with subject_code={subject_code}")
            if teacher_code and teacher_record is None:
                row_errors.append(f"Teacher not found with teacher_code={teacher_code}")
            if room_number is not None and room_record is None:
                row_errors.append(f"Room not found with room_number={room_number}")

            item = UploadTeachingCalenderItem(
                subject_id=subject_record.id if subject_record else None,
                subject_code=subject_code,
                subject_name=subject_name or (subject_record.name if subject_record else None),
                teacher_id=teacher_record.id if teacher_record else None,
                teacher_code=teacher_code,
                teacher_name=teacher_name or (teacher_record.name if teacher_record else None),
                weeekday=weekday_number or 0,
                room_id=room_record.id if room_record else None,
                room_number=room_number,
                lesson_periods=lesson_periods or "",
                study_weeks=study_weeks or "",
            )

            if row_errors:
                invalid_rows.append(
                    UploadTeachingCalenderInvalidRow(
                        **item.model_dump(),
                        row=row_index,
                        errors=row_errors,
                    )
                )
                continue

            parsed_rows.append(item)

        return UploadTeachingCalenderResponse(
            file_information=UploadTeachingCalenderFileInfo(
                file_name=filename,
                headers=expected_headers,
                header_row=header_row_index,
                total_rows=max(len(rows) - (header_row_index + 1), 0),
                valid_rows_count=len(parsed_rows),
                invalid_rows_count=len(invalid_rows),
            ),
            schedules=parsed_rows,
            invalid_schedules=invalid_rows,
        )

    @staticmethod
    def import_calender(
        *, session: Session, request: Request, calender: ImportTeachingCalenderInput 
    ) -> ImportTeachingCalenderResponse:
        start_date = calender.period.start_date
        end_date = calender.period.end_date

        if end_date < start_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="end_date must be greater than or equal to start_date.",
            )

        class_record = session.exec(
            select(Classes).where(Classes.class_code == calender.class_code)
        ).first()
        if not class_record:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Class not found with class_code={calender.class_code}",
            )

        imported_items: list[ImportTeachingCalenderImportedItem] = []
        pending_learning_schedules: list[dict] = []

        try:
            for row_index, schedule in enumerate(calender.schedules, start=1):
                if not schedule.subject_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing subject_id at row {row_index}. Please upload/resolve file before import.",
                    )
                if not schedule.teacher_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing teacher_id at row {row_index}. Please upload/resolve file before import.",
                    )
                if not schedule.room_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing room_id at row {row_index}. Please upload/resolve file before import.",
                    )

                period_blocks = TeachingScheduleServices._parse_lesson_periods(
                    schedule.lesson_periods
                )
                study_week_indexes = TeachingScheduleServices._parse_study_week_indexes(
                    schedule.study_weeks
                )
                weekday_number = schedule.weeekday
                TeachingScheduleServices._validate_weekday_number(weekday_number)

                for week_index in study_week_indexes:
                    matched_date = TeachingScheduleServices._resolve_date_for_week(
                        period_start_date=start_date,
                        week_index=week_index,
                        weekday_number=weekday_number,
                    )
                    if matched_date > end_date:
                        continue

                    for start_period, end_period in period_blocks:
                        pending_learning_schedules.append(
                            {
                                "row": row_index,
                                "class_id": class_record.id,
                                "class_code": class_record.class_code,
                                "subject_id": schedule.subject_id,
                                "subject_code": schedule.subject_code,
                                "subject_name": schedule.subject_name,
                                "teacher_id": schedule.teacher_id,
                                "room_id": schedule.room_id,
                                "date": matched_date,
                                "start_period": start_period,
                                "end_period": end_period,
                            }
                        )

            subject_ids = list({item["subject_id"] for item in pending_learning_schedules})
            existing_learning_schedules = session.exec(
                select(LearningSchedules).where(
                    LearningSchedules.class_id == class_record.id,
                    LearningSchedules.subject_id.in_(subject_ids),
                    LearningSchedules.date >= start_date,
                    LearningSchedules.date <= end_date,
                    or_(
                        LearningSchedules.status == StatusEnum.ACTIVE,
                        LearningSchedules.status.is_(None),
                    ),
                )
            ).all() if subject_ids else []

            existing_group_map: dict[tuple[uuid.UUID, datetime], list[LearningSchedules]] = {}
            for existing_item in existing_learning_schedules:
                group_key = (existing_item.subject_id, existing_item.date)
                existing_group_map.setdefault(group_key, []).append(existing_item)

            pending_group_map: dict[tuple[uuid.UUID, datetime], list[dict]] = {}
            for item in pending_learning_schedules:
                group_key = (item["subject_id"], item["date"])
                pending_group_map.setdefault(group_key, []).append(item)

            for group_key, pending_items in pending_group_map.items():
                pending_items.sort(key=lambda item: (item["start_period"], item["end_period"]))

                for pending_item in pending_items:
                    for existing_item in existing_group_map.get(group_key, []):
                        if (
                            existing_item.start_period <= pending_item["end_period"]
                            and existing_item.end_period >= pending_item["start_period"]
                        ):
                            raise HTTPException(
                                status_code=status.HTTP_400_BAD_REQUEST,
                                detail=TeachingScheduleServices._build_duplicate_schedule_message(
                                    class_code=pending_item["class_code"],
                                    subject_code=pending_item["subject_code"],
                                    subject_name=pending_item["subject_name"],
                                    date_value=pending_item["date"],
                                    start_period=existing_item.start_period,
                                    end_period=existing_item.end_period,
                                ),
                            )

                for index, pending_item in enumerate(pending_items[:-1]):
                    next_item = pending_items[index + 1]
                    if pending_item["end_period"] >= next_item["start_period"]:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=TeachingScheduleServices._build_duplicate_schedule_message(
                                class_code=pending_item["class_code"],
                                subject_code=pending_item["subject_code"],
                                subject_name=pending_item["subject_name"],
                                date_value=pending_item["date"],
                                start_period=next_item["start_period"],
                                end_period=next_item["end_period"],
                            ),
                        )

            for item in pending_learning_schedules:
                learning_schedule = LearningSchedules(
                    class_id=item["class_id"],
                    subject_id=item["subject_id"],
                    date=item["date"],
                    start_period=item["start_period"],
                    end_period=item["end_period"],
                    room_id=item["room_id"],
                    schedule_type=None,
                    status=StatusEnum.ACTIVE,
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                )
                session.add(learning_schedule)
                session.flush()

                teaching_schedule = TeachingSchedules(
                    teacher_id=item["teacher_id"],
                    learning_schedule_id=learning_schedule.id,
                    status=StatusEnum.ACTIVE,
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                )
                session.add(teaching_schedule)
                session.flush()

                imported_items.append(
                    ImportTeachingCalenderImportedItem(
                        row=item["row"],
                        date=item["date"],
                        learning_schedule_id=learning_schedule.id,
                        teaching_schedule_id=teaching_schedule.id,
                    )
                )

            session.commit()
        except HTTPException:
            session.rollback()
            raise
        except Exception:
            session.rollback()
            raise

        return ImportTeachingCalenderResponse(items=imported_items)
