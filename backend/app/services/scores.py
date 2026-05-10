import uuid
from io import BytesIO
import re
from datetime import datetime
from fastapi import HTTPException, Request
from fastapi import UploadFile
from sqlmodel import Session, func, or_, select
from starlette import status
from typing import List
from openpyxl import load_workbook

from app.models.models import (
    AcademicTerms,
    Classes,
    ScoreComponents,
    Scores,
    StudentClass,
    Students,
    Subjects,
)
from app.models.schemas.scores.score_schemas import (
    ScoreByClassSubjectParams,
    ScoreByClassSubjectResponse,
    ScoreAggregationBucket,
    ScorePointItem,
    StudentAndGpaListItem,
    StudentAndGpaListResponse,
    ScoresPublic,
    ScoresCreate,
    ScoresUpdate,
    ScoresDeleteResponse,
    ScoreBulkCreatePayload,
    ScoreBulkCreateResponse,
    ScoreBulkUpdateItem,
    ScoreBulkUpdatePayload,
    ScoreBulkUpdateResponse,
    ScoreBulkStatusUpdatePayload,
    ScoreBulkStatusUpdateResponse,
    StudentScoreByClassSubjectItem,
    StudentScoreByStudentResponse,
    StudentScoreFilterParams,
    StudentAndGpaResponse,
    StudentGpaClassInfo,
    StudentGpaSummary,
    StudentScoreItemResponse,
    StudentScoreComponentResponse,
    StudentInfoScoreResponse,
    StudentScoresPayload,
    ScoreFileData,
    ScoreFileDataResponse,
    ScoreFileInfo,
    ScoreFileInvalidRow,
    ScoreImportItem,
    ScoreImportListPayload,
    ScoreImportListResponse,
    ScoreImportCreatedItem,
)
from app.enums.status import StatusEnum
from app.enums.grade import (
    GRADE_SCALE_THRESHOLDS,
    GradeRankEnum,
    ScoreComponentTypeEnum,
    ScoreTypeEnum,
    GradeScaleEnum
    , MIDDLE_COMPONENT_TYPE_ALIASES, FINAL_COMPONENT_TYPE_ALIASES, OTHER_COMPONENT_TYPE_ALIASES
)
from app.services.common import to_clean_text


class ScoresServices:
    @staticmethod
    def _resolve_score_component_id(
        *,
        session: Session,
        component_type: str | None = None,
        score_component_id: uuid.UUID | None = None,
    ) -> uuid.UUID:
        if score_component_id is not None:
            component = session.get(ScoreComponents, score_component_id)
            if component is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Score component not found with id={score_component_id}",
                )
            return component.id

        if not component_type:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="score_component_id or component_type is required.",
            )

        component = session.exec(
            select(ScoreComponents).where(func.upper(ScoreComponents.component_type) == component_type.upper())
        ).first()
        if component is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Score component not found for component_type={component_type}",
            )
        return component.id

    @staticmethod
    def _resolve_student_id(
        *,
        session: Session,
        student_id: uuid.UUID | None,
        student_code: str | None,
    ) -> uuid.UUID:
        if student_id is not None:
            return student_id
        if not student_code:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="student_id or student_code is required.",
            )
        student = session.exec(
            select(Students).where(Students.student_code == student_code)
        ).first()
        if student is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Student not found with student_code={student_code}",
            )
        return student.id

    @staticmethod
    def _cell_text(cell: object, number_format: str | None = None) -> str | None:
        if cell is None or cell == "":
            return None
        if isinstance(cell, float) and cell.is_integer():
            cell = int(cell)
        if isinstance(cell, int) and number_format:
            zero_match = re.fullmatch(r"0+", number_format.strip())
            if zero_match:
                return str(cell).zfill(len(zero_match.group(0)))
        return to_clean_text(cell)

    @staticmethod
    def _metadata_value(cell: object, number_format: str | None = None) -> str | None:
        text = ScoresServices._cell_text(cell, number_format)
        if not text:
            return None
        if ":" in text:
            return to_clean_text(text.split(":", maxsplit=1)[1])
        return text

    @staticmethod
    def _parse_score_value(raw_value: object) -> float | None:
        if raw_value is None or raw_value == "":
            return None
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        text_value = to_clean_text(raw_value)
        if text_value is None:
            return None
        return float(text_value.replace(",", "."))

    @staticmethod
    def _extract_metadata(worksheet) -> dict[str, str | int | None]:
        metadata: dict[str, str | int | None] = {
            "class_code": None,
            "academic_year": None,
            "semester": None,
            "subject_name": None,
            "subject_code": None,
            "attempt": None,
        }

        # The template places metadata in fixed cells.
        metadata["class_code"] = ScoresServices._metadata_value(
            worksheet["A5"].value,
            worksheet["A5"].number_format,
        )
        metadata["academic_year"] = ScoresServices._metadata_value(
            worksheet["A6"].value,
            worksheet["A6"].number_format,
        )
        semester_text = ScoresServices._metadata_value(
            worksheet["A7"].value,
            worksheet["A7"].number_format,
        )
        if semester_text and semester_text.isdigit():
            metadata["semester"] = int(semester_text)
        metadata["subject_name"] = ScoresServices._metadata_value(
            worksheet["A8"].value,
            worksheet["A8"].number_format,
        )
        metadata["subject_code"] = ScoresServices._metadata_value(
            worksheet["A9"].value,
            worksheet["A9"].number_format,
        )
        attempt_text = ScoresServices._metadata_value(
            worksheet["A10"].value,
            worksheet["A10"].number_format,
        )
        if attempt_text and attempt_text.isdigit():
            metadata["attempt"] = int(attempt_text)

        return metadata

    @staticmethod
    async def upload_file_score(
        *,
        session: Session,
        file: UploadFile,
    ) -> ScoreFileDataResponse:
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows())
        if len(rows) < 12:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not contain the expected score header row.",
            )

        metadata = ScoresServices._extract_metadata(worksheet)
        academic_term_id = None
        if metadata["academic_year"] and metadata["semester"] is not None:
            academic_term = session.exec(
                select(AcademicTerms).where(
                    AcademicTerms.academic_year == metadata["academic_year"],
                    AcademicTerms.semester == metadata["semester"],
                )
            ).first()
            if academic_term is not None:
                academic_term_id = academic_term.id

        subject_id = None
        if metadata["subject_code"]:
            subject = session.exec(
                select(Subjects).where(Subjects.subject_code == metadata["subject_code"])
            ).first()
            if subject is not None:
                subject_id = subject.id

        header_row_index = 11
        data_start_row_index = header_row_index + 1
        expected_headers = [
            "STT",
            "Lớp",
            "MSV",
            "Họ và đệm",
            "Tên",
            "D1",
            "D2",
            "THI",
            "TBM",
            "Ghi chú",
        ]

        parsed_rows: list[ScoreFileData] = []
        invalid_rows: list[ScoreFileInvalidRow] = []
        data_started = False
        empty_row_count = 0

        for row_index, row_values in enumerate(rows[data_start_row_index:], start=data_start_row_index + 1):
            if row_values is None:
                if data_started:
                    empty_row_count += 1
                    if empty_row_count >= 1:
                        break
                continue

            def cell(col_index: int) -> object | None:
                if col_index >= len(row_values):
                    return None
                return row_values[col_index].value

            def cell_text(col_index: int) -> str | None:
                if col_index >= len(row_values):
                    return None
                current_cell = row_values[col_index]
                return ScoresServices._cell_text(current_cell.value, current_cell.number_format)

            stt_raw = cell(0)
            class_code = cell_text(1)
            student_code = cell_text(2)
            family_name = cell_text(3)
            given_name = cell_text(4)
            d1_raw = cell(5)
            d2_raw = cell(6)
            thi_raw = cell(7)
            tbm_raw = cell(8)
            note = cell_text(9)

            if all(
                value in (None, "")
                for value in (
                    stt_raw,
                    class_code,
                    student_code,
                    family_name,
                    given_name,
                    d1_raw,
                    d2_raw,
                    thi_raw,
                    tbm_raw,
                    note,
                )
            ):
                if data_started:
                    empty_row_count += 1
                    if empty_row_count >= 1:
                        break
                continue

            data_started = True
            empty_row_count = 0

            row_errors: list[str] = []
            student_id = None

            stt: int | None = None
            if stt_raw not in (None, ""):
                try:
                    stt = int(float(stt_raw))
                except (TypeError, ValueError):
                    row_errors.append("STT must be a number.")

            d1 = None
            d2 = None
            thi = None
            tbm = None
            try:
                d1 = ScoresServices._parse_score_value(d1_raw)
                d2 = ScoresServices._parse_score_value(d2_raw)
                thi = ScoresServices._parse_score_value(thi_raw)
                tbm = ScoresServices._parse_score_value(tbm_raw)
            except ValueError:
                row_errors.append("Score values must be numeric.")

            if not class_code:
                row_errors.append("Class Code is required.")
            if not student_code:
                row_errors.append("Student Code is required.")
            if not family_name and not given_name:
                row_errors.append("Student name is required.")
            if d1 is None:
                row_errors.append("D1 is required.")
            if d2 is None:
                row_errors.append("D2 is required.")
            if thi is None:
                row_errors.append("THI is required.")

            if class_code:
                class_record = session.exec(
                    select(Classes).where(Classes.class_code == class_code)
                ).first()
                if class_record is None:
                    row_errors.append(f"Class not found with class_code={class_code}")

            if student_code:
                student_record = session.exec(
                    select(Students).where(Students.student_code == student_code)
                ).first()
                if student_record is None:
                    row_errors.append(f"Student not found with student_code={student_code}")
                else:
                    student_id = student_record.id

            if tbm is not None and (tbm < 0 or tbm > 10):
                row_errors.append("TBM must be between 0 and 10.")

            if row_errors:
                invalid_rows.append(
                    ScoreFileInvalidRow(
                        row=row_index,
                        stt=stt,
                        class_code=class_code,
                        student_code=student_code,
                        student_id=student_id,
                        student_name=" ".join(part for part in [family_name, given_name] if part) or None,
                        family_name=family_name,
                        given_name=given_name,
                        d1=d1,
                        d2=d2,
                        thi=thi,
                        tbm=tbm,
                        note=note,
                        errors=row_errors,
                    )
                )
                continue

            parsed_rows.append(
                ScoreFileData(
                    row=row_index,
                    stt=stt,
                    class_code=class_code,
                    student_code=student_code,
                    student_id=student_id,
                    student_name=" ".join(part for part in [family_name, given_name] if part) or None,
                    family_name=family_name,
                    given_name=given_name,
                    d1=d1,
                    d2=d2,
                    thi=thi,
                    tbm=tbm,
                    note=note,
                )
            )

        return ScoreFileDataResponse(
            file_information=ScoreFileInfo(
                file_name=filename,
                headers=expected_headers,
                header_row=header_row_index + 1,
                total_rows=max(len(rows) - (header_row_index + 1), 0),
                valid_rows_count=len(parsed_rows),
                invalid_rows_count=len(invalid_rows),
                class_code=metadata["class_code"] if isinstance(metadata["class_code"], str) else None,
                academic_year=metadata["academic_year"] if isinstance(metadata["academic_year"], str) else None,
                semester=metadata["semester"] if isinstance(metadata["semester"], int) else None,
                academic_term_id=academic_term_id,
                subject_name=metadata["subject_name"] if isinstance(metadata["subject_name"], str) else None,
                subject_code=metadata["subject_code"] if isinstance(metadata["subject_code"], str) else None,
                subject_id=subject_id,
                attempt=metadata["attempt"] if isinstance(metadata["attempt"], int) else None,
            ),
            scores=parsed_rows,
            invalid_scores=invalid_rows,
        )

    @staticmethod
    def import_score_list(
        *,
        session: Session,
        payload: ScoreImportListPayload,
    ) -> ScoreImportListResponse:
        if payload.attempt < 1:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="attempt must be greater than or equal to 1.",
            )

        if not session.get(AcademicTerms, payload.academic_term_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Academic term not found with id={payload.academic_term_id}",
            )
        if not session.get(Subjects, payload.subject_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Subject not found with id={payload.subject_id}",
            )

        middle_component_id = ScoresServices._resolve_score_component_id(
            session=session,
            component_type=ScoreComponentTypeEnum.MIDDLE.value.capitalize(),
        )
        final_component_id = ScoresServices._resolve_score_component_id(
            session=session,
            component_type=ScoreComponentTypeEnum.FINAL.value.capitalize(),
        )

        imported_items: list[ScoreImportCreatedItem] = []
        for item in payload.scores:
            resolved_student_id = ScoresServices._resolve_student_id(
                session=session,
                student_id=item.student_id,
                student_code=item.student_code,
            )

            resolved_scores: list[Scores] = []
            point_values = [
                (item.score_1, middle_component_id),
                (item.score_2, middle_component_id),
                (item.score_exam, final_component_id),
            ]

            for score_value, component_id in point_values:
                if score_value is None:
                    continue
                new_score = Scores(
                    student_id=resolved_student_id,
                    subject_id=payload.subject_id,
                    score_component_id=component_id,
                    academic_term_id=payload.academic_term_id,
                    score=score_value,
                    attempt=payload.attempt,
                    score_type=ScoreTypeEnum.OFFICIAL.value.capitalize()
                    if payload.attempt == 1
                    else ScoreTypeEnum.RETAKE.value.capitalize(),
                    status="pending",
                )
                session.add(new_score)
                session.flush()
                resolved_scores.append(new_score)

            imported_items.append(
                ScoreImportCreatedItem(
                    student_id=resolved_student_id,
                    subject_id=payload.subject_id,
                    academic_term_id=payload.academic_term_id,
                    score_ids=[score.id for score in resolved_scores if score.id is not None],
                )
            )

        session.commit()

        return ScoreImportListResponse(
            items=imported_items,
            total=len(imported_items),
        )

    @staticmethod
    def _build_student_gpa_payload(
        *, session: Session, student: Students
    ) -> StudentAndGpaListItem:
        from app.services.students import StudentServices

        class_id = StudentServices._get_primary_class_id(
            session=session,
            student_id=student.id,
        )

        class_code = None
        class_name = None
        if class_id is not None:
            class_ = session.get(Classes, class_id)
            if class_ is not None:
                class_code = class_.class_code
                class_name = class_.class_name

        # reuse the same GPA logic as the single-student endpoint
        single = ScoresServices.get_student_and_gpa(session=session, student_id=student.id)
        return StudentAndGpaListItem(
            student_info=single.student_info,
            class_info=StudentGpaClassInfo(
                class_id=class_id,
                class_code=class_code,
                class_name=class_name,
            ),
            gpa=single.gpa,
        )

    @staticmethod
    def get_student_and_gpa(
        *, session: Session, student_id: uuid.UUID
    ) -> StudentAndGpaResponse:
        student = session.get(Students, student_id)
        if not student:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
            )

        from app.services.students import StudentServices

        class_id = StudentServices._get_primary_class_id(
            session=session,
            student_id=student.id,
        )

        class_code = None
        class_name = None
        if class_id is not None:
            class_ = session.get(Classes, class_id)
            if class_ is not None:
                class_code = class_.class_code
                class_name = class_.class_name

        def normalize_score_text(value: str) -> str:
            return value.upper().strip()

        def is_retake_score(score_type: str, attempt: int) -> bool:
            score_type_normalized = normalize_score_text(score_type or "")
            if score_type_normalized == ScoreTypeEnum.RETAKE.value.upper():
                return True
            if score_type_normalized == ScoreTypeEnum.OFFICIAL.value.upper():
                return False
            return attempt > 1

        def is_midterm_component(component_type: str) -> bool:
            return normalize_score_text(component_type) in MIDDLE_COMPONENT_TYPE_ALIASES

        def is_final_component(component_type: str) -> bool:
            return normalize_score_text(component_type) in FINAL_COMPONENT_TYPE_ALIASES

        def is_other_component(component_type: str) -> bool:
            return normalize_score_text(component_type) in OTHER_COMPONENT_TYPE_ALIASES

        def to_normalized_weight(weight: float) -> float:
            if weight > 1:
                return weight / 100
            if weight < 0:
                return 0
            return weight

        def score10_to_scale(value: float) -> tuple[float, str]:
            for min_score, avg4_value, letter in GRADE_SCALE_THRESHOLDS:
                if value >= min_score:
                    return avg4_value, letter.value
            return 0.0, GradeScaleEnum.F

        score_rows = session.exec(
            select(
                Scores,
                ScoreComponents.component_type.label("component_type"),
                ScoreComponents.weight.label("component_weight"),
                AcademicTerms.academic_year.label("academic_year"),
                AcademicTerms.semester.label("semester"),
            )
            .join(ScoreComponents, ScoreComponents.id == Scores.score_component_id)
            .join(AcademicTerms, AcademicTerms.id == Scores.academic_term_id)
            .where(Scores.student_id == student.id)
            .order_by(Scores.created_at.desc())
        ).all()

        aggregated: dict[str, ScoreAggregationBucket] = {}
        for score, component_type, component_weight, academic_year, semester in score_rows:
            if score.status is not None and normalize_score_text(str(score.status)) == "INACTIVE":
                continue

            key = f"{score.subject_id}-{score.academic_term_id}"
            bucket = aggregated.setdefault(
                key,
                ScoreAggregationBucket(
                    subject_id=score.subject_id,
                    academic_term_id=score.academic_term_id,
                    academic_year=academic_year,
                    semester=semester,
                ),
            )
            bucket.points.append(
                ScorePointItem(
                    score=score.score,
                    weight=component_weight,
                    attempt=score.attempt,
                    score_type=score.score_type,
                    component_type=component_type,
                )
            )

        has_any_score = False
        studied_credits = 0
        accumulated_credits = 0
        gpa_weighted_10 = 0.0
        gpa_weighted_4 = 0.0
        gpa_weight_sum = 0.0

        for bucket in aggregated.values():
            subject = session.get(Subjects, bucket.subject_id)
            if subject is None:
                continue

            studied_credits += subject.credit or 0

            points = bucket.points
            official_mid: list[ScorePointItem] = []
            official_final: ScorePointItem | None = None
            retake_mid: list[ScorePointItem] = []
            retake_final: ScorePointItem | None = None

            for point in points:
                is_retake = is_retake_score(point.score_type, point.attempt)
                component_type = point.component_type
                if is_retake:
                    if is_final_component(component_type):
                        retake_final = point
                    elif is_midterm_component(component_type):
                        retake_mid.append(point)
                    elif is_other_component(component_type) and len(retake_mid) < 2:
                        retake_mid.append(point)
                    else:
                        retake_final = point
                else:
                    if is_final_component(component_type):
                        official_final = point
                    elif is_midterm_component(component_type):
                        official_mid.append(point)
                    elif is_other_component(component_type) and len(official_mid) < 2:
                        official_mid.append(point)
                    else:
                        official_final = point

            selected_mid1 = retake_mid[0] if len(retake_mid) > 0 else (official_mid[0] if len(official_mid) > 0 else None)
            selected_mid2 = retake_mid[1] if len(retake_mid) > 1 else (official_mid[1] if len(official_mid) > 1 else None)
            selected_final = retake_final or official_final
            selected_points = [
                p
                for p in [selected_mid1, selected_mid2, selected_final]
                if p is not None and p.score is not None
            ]
            if not selected_points:
                continue
            has_any_score = True

            normalized_weight_sum = sum(to_normalized_weight(p.weight) for p in selected_points)
            if normalized_weight_sum > 0:
                avg10 = sum(
                    p.score * to_normalized_weight(p.weight)
                    for p in selected_points
                ) / normalized_weight_sum
            else:
                avg10 = sum(p.score for p in selected_points) / len(selected_points)

            avg10 = round(avg10, 2)
            avg4, letter = score10_to_scale(avg10)

            gpa_weighted_10 += avg10 * (subject.credit or 0)
            gpa_weighted_4 += avg4 * (subject.credit or 0)
            gpa_weight_sum += subject.credit or 0

            if avg10 >= 4.0:
                accumulated_credits += subject.credit or 0

        gpa10 = round(gpa_weighted_10 / gpa_weight_sum, 2) if gpa_weight_sum > 0 else 0.0
        gpa4 = round(gpa_weighted_4 / gpa_weight_sum, 2) if gpa_weight_sum > 0 else 0.0

        def classify_gpa4(value: float) -> str:
            if not has_any_score:
                return GradeRankEnum.NOT_RANKED
            if value >= 3.6:
                return GradeRankEnum.EXCELLENT
            if value >= 3.2:
                return GradeRankEnum.GOOD
            if value >= 2.5:
                return GradeRankEnum.FAIR
            if value >= 2.0:
                return GradeRankEnum.AVERAGE
            return GradeRankEnum.POOR

        def classify_gpa10(value: float) -> str:
            if not has_any_score:
                return GradeRankEnum.NOT_RANKED
            if value >= 9.0:
                return GradeRankEnum.EXCELLENT
            if value >= 8.0:
                return GradeRankEnum.GOOD
            if value >= 6.5:
                return GradeRankEnum.FAIR
            if value >= 5.0:
                return GradeRankEnum.AVERAGE
            return GradeRankEnum.POOR

        return StudentAndGpaResponse(
            student_info=StudentInfoScoreResponse(
                id=student.id,
                student_code=student.student_code,
                name=student.name,
                email=student.email,
                phone=student.phone,
            ),
            class_info=StudentGpaClassInfo(
                class_id=class_id,
                class_code=class_code,
                class_name=class_name,
            ),
            gpa=StudentGpaSummary(
                grade4=classify_gpa4(gpa4),
                grade10=classify_gpa10(gpa10),
                gpa4=gpa4,
                accumulated_gpa4=gpa4,
                accumulated_gpa10=gpa10,
                accumulated_credits=accumulated_credits,
                studied_credits=studied_credits,
            ),
        )

    @staticmethod
    def get_students_and_gpa(
        *, session: Session, query
    ) -> StudentAndGpaListResponse:
        statement = select(Students)
        conditions = []
        if query.status:
            conditions.append(Students.status == query.status)

        if query.class_id:
            student_ids_by_class = session.exec(
                select(StudentClass.student_id).where(
                    StudentClass.class_id == query.class_id,
                    or_(
                        StudentClass.status != StatusEnum.INACTIVE,
                        StudentClass.status.is_(None),
                    ),
                )
            ).all()
            filtered_student_ids = [
                student_id for student_id in student_ids_by_class if student_id is not None
            ]
            if not filtered_student_ids:
                return StudentAndGpaListResponse(data=[], total=0)
            conditions.append(Students.id.in_(filtered_student_ids))

        if query.search:
            conditions.append(
                or_(
                    Students.student_code.ilike(f"%{query.search}%"),
                    Students.name.ilike(f"%{query.search}%"),
                )
            )

        if conditions:
            statement = statement.where(*conditions)

        count_stmt = select(func.count()).select_from(Students)
        if conditions:
            count_stmt = count_stmt.where(*conditions)
        total = session.exec(count_stmt).one()

        student_rows = session.exec(
            statement.order_by(Students.created_at.desc()).offset(query.skip).limit(query.limit)
        ).all()

        data = [
            ScoresServices._build_student_gpa_payload(session=session, student=student)
            for student in student_rows
        ]
        return StudentAndGpaListResponse(data=data, total=total)

    @staticmethod
    def get_all(
        *,
        session: Session,
    ) -> List[ScoresPublic]:
        scores = session.exec(select(Scores)).all()
        return scores

    @staticmethod
    def get_by_id(
        *, session: Session, score_id: uuid.UUID, request: Request
    ) -> ScoresPublic:
        score = session.get(Scores, score_id)
        if not score:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Score does not exist"
            )
        return ScoresPublic.model_validate(score)

    @staticmethod
    def get_by_student(
        *,
        session: Session,
        student_id: uuid.UUID,
        query: StudentScoreFilterParams,
    ) -> StudentScoreByStudentResponse:
        student = session.get(Students, student_id)
        if not student:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
            )

        statement = (
            select(
                Scores,
                ScoreComponents.id.label("component_id"),
                ScoreComponents.component_type.label("component_type"),
                ScoreComponents.weight.label("component_weight"),
                ScoreComponents.description.label("component_description"),
                Subjects.subject_code.label("subject_code"),
                Subjects.name.label("subject_name"),
                Subjects.credit.label("subject_credit"),
                AcademicTerms.academic_year.label("academic_year"),
                AcademicTerms.semester.label("semester"),
            )
            .join(ScoreComponents, ScoreComponents.id == Scores.score_component_id)
            .join(Subjects, Subjects.id == Scores.subject_id)
            .join(AcademicTerms, AcademicTerms.id == Scores.academic_term_id)
            .where(Scores.student_id == student_id)
        )

        if query.academic_term_id:
            statement = statement.where(Scores.academic_term_id == query.academic_term_id)

        if query.subject_id:
            statement = statement.where(Scores.subject_id == query.subject_id)

        rows = session.exec(statement.order_by(Scores.created_at.desc())).all()

        score_items: list[StudentScoreItemResponse] = []
        for (
            score,
            component_id,
            component_type,
            component_weight,
            component_description,
            subject_code,
            subject_name,
            subject_credit,
            academic_year,
            semester,
        ) in rows:
            score_items.append(
                StudentScoreItemResponse(
                    id=score.id,
                    subject_id=score.subject_id,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    subject_credit=subject_credit,
                    academic_term_id=score.academic_term_id,
                    academic_year=academic_year,
                    semester=semester,
                    score=score.score,
                    attempt=score.attempt,
                    score_type=score.score_type,
                    status=score.status,
                    created_at=score.created_at,
                    updated_at=score.updated_at,
                    score_component=StudentScoreComponentResponse(
                        id=component_id,
                        component_type=component_type,
                        weight=component_weight,
                        description=component_description,
                    ),
                )
            )

        return StudentScoreByStudentResponse(
            student_info=StudentInfoScoreResponse(
                id=student.id,
                student_code=student.student_code,
                name=student.name,
                email=student.email,
                phone=student.phone,
            ),
            scores=StudentScoresPayload(
                items=score_items,
                total=len(score_items),
            ),
        )

    @staticmethod
    def get_by_class_subject(
        *,
        session: Session,
        query: ScoreByClassSubjectParams,
    ) -> ScoreByClassSubjectResponse:
        class_ = session.get(Classes, query.class_id)
        if not class_:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Class not found"
            )

        subject = session.get(Subjects, query.subject_id)
        if not subject:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Subject not found"
            )

        student_ids_rows = session.exec(
            select(StudentClass.student_id).where(
                StudentClass.class_id == query.class_id,
                or_(
                    StudentClass.status != StatusEnum.INACTIVE,
                    StudentClass.status.is_(None),
                ),
            )
        ).all()

        student_ids: list[uuid.UUID] = []
        for student_id in student_ids_rows:
            if student_id is None:
                continue
            if student_id not in student_ids:
                student_ids.append(student_id)

        if not student_ids:
            return ScoreByClassSubjectResponse(
                class_id=class_.id,
                class_name=class_.class_name,
                subject_id=subject.id,
                subject_name=subject.name,
                students=[],
                total_students=0,
            )

        students = session.exec(
            select(Students)
            .where(Students.id.in_(student_ids))
            .order_by(Students.name.asc())
        ).all()

        score_rows = session.exec(
            select(
                Scores,
                ScoreComponents.id.label("component_id"),
                ScoreComponents.component_type.label("component_type"),
                ScoreComponents.weight.label("component_weight"),
                ScoreComponents.description.label("component_description"),
                AcademicTerms.academic_year.label("academic_year"),
                AcademicTerms.semester.label("semester"),
            )
            .join(ScoreComponents, ScoreComponents.id == Scores.score_component_id)
            .join(AcademicTerms, AcademicTerms.id == Scores.academic_term_id)
            .where(
                Scores.student_id.in_(student_ids),
                Scores.subject_id == query.subject_id,
                or_(
                    Scores.status != StatusEnum.INACTIVE,
                    Scores.status.is_(None),
                ),
            )
            .order_by(Scores.student_id, Scores.created_at.desc())
        ).all()

        scores_by_student: dict[uuid.UUID, list[StudentScoreItemResponse]] = {}
        for (
            score,
            component_id,
            component_type,
            component_weight,
            component_description,
            academic_year,
            semester,
        ) in score_rows:
            if score.student_id is None:
                continue

            scores_by_student.setdefault(score.student_id, []).append(
                StudentScoreItemResponse(
                    id=score.id,
                    subject_id=score.subject_id,
                    subject_code=subject.subject_code,
                    subject_name=subject.name,
                    subject_credit=subject.credit,
                    academic_term_id=score.academic_term_id,
                    academic_year=academic_year,
                    semester=semester,
                    score=score.score,
                    attempt=score.attempt,
                    score_type=score.score_type,
                    status=score.status,
                    created_at=score.created_at,
                    updated_at=score.updated_at,
                    score_component=StudentScoreComponentResponse(
                        id=component_id,
                        component_type=component_type,
                        weight=component_weight,
                        description=component_description,
                    ),
                )
            )

        student_items = [
            StudentScoreByClassSubjectItem(
                student_info=StudentInfoScoreResponse(
                    id=student.id,
                    student_code=student.student_code,
                    name=student.name,
                    email=student.email,
                    phone=student.phone,
                ),
                scores=scores_by_student.get(student.id, []),
            )
            for student in students
        ]

        return ScoreByClassSubjectResponse(
            class_id=class_.id,
            class_name=class_.class_name,
            subject_id=subject.id,
            subject_name=subject.name,
            students=student_items,
            total_students=len(student_items),
        )

    @staticmethod
    def create(
        *,
        session: Session,
        score: ScoresCreate,
    ) -> ScoresPublic:
        resolved_score_component_id = score.score_component_id
        if resolved_score_component_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="score_component_id is required.",
            )

        existing = session.exec(
            select(Scores).where(
                Scores.student_id == score.student_id,
                Scores.score_component_id == resolved_score_component_id,
                Scores.attempt == score.attempt,
            )
        ).first()

        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Score already exists.",
            )

        new_score = Scores(**score.dict(exclude={"score_component_id"}), score_component_id=resolved_score_component_id)
        session.add(new_score)
        session.commit()
        session.refresh(new_score)

        return new_score

    @staticmethod
    def bulk_create(
        *,
        session: Session,
        payload: ScoreBulkCreatePayload,
    ) -> ScoreBulkCreateResponse:
        if not payload.scores:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="scores must not be empty.",
            )

        seen_keys: set[tuple[uuid.UUID, uuid.UUID, int]] = set()
        for item in payload.scores:
            resolved_score_component_id = item.score_component_id
            if resolved_score_component_id is None:
                resolved_score_component_id = ScoresServices._resolve_score_component_id(
                    session=session,
                    component_type=getattr(item, "component_type", None),
                )

            key = (item.student_id, resolved_score_component_id, item.attempt)
            if key in seen_keys:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Duplicate score item in request payload.",
                )
            seen_keys.add(key)

            existing = session.exec(
                select(Scores).where(
                    Scores.student_id == item.student_id,
                    Scores.score_component_id == resolved_score_component_id,
                    Scores.attempt == item.attempt,
                )
            ).first()
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Score already exists.",
                )

        created_items: list[Scores] = []
        for item in payload.scores:
            resolved_score_component_id = item.score_component_id
            if resolved_score_component_id is None:
                resolved_score_component_id = ScoresServices._resolve_score_component_id(
                    session=session,
                    component_type=getattr(item, "component_type", None),
                )

            new_score = Scores(
                **item.model_dump(exclude={"component_type", "score_component_id"}),
                score_component_id=resolved_score_component_id,
            )
            session.add(new_score)
            session.flush()
            created_items.append(new_score)

        session.commit()
        for item in created_items:
            session.refresh(item)

        return ScoreBulkCreateResponse(items=created_items, total=len(created_items))

    @staticmethod
    def bulk_update(
        *,
        session: Session,
        payload: ScoreBulkUpdatePayload,
    ) -> ScoreBulkUpdateResponse:
        if not payload.scores:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="scores must not be empty.",
            )

        seen_ids: set[uuid.UUID] = set()
        for item in payload.scores:
            if item.id in seen_ids:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Duplicate score id in request payload.",
                )
            seen_ids.add(item.id)

        updated_items: list[Scores] = []
        for item in payload.scores:
            score = session.get(Scores, item.id)
            if not score:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Score not found with id={item.id}",
                )

            update_data = item.model_dump(exclude_unset=True)
            component_type = update_data.pop("component_type", None)
            score_component_id = update_data.get("score_component_id")
            if score_component_id is None and component_type is not None:
                update_data["score_component_id"] = ScoresServices._resolve_score_component_id(
                    session=session,
                    component_type=component_type,
                )
            for field, value in update_data.items():
                if field == "id":
                    continue
                setattr(score, field, value)

            if score.score_type is None:
                score.score_type = ScoreTypeEnum.OFFICIAL.value.capitalize()

            updated_items.append(score)

        session.commit()
        for item in updated_items:
            session.refresh(item)

        return ScoreBulkUpdateResponse(items=updated_items, total=len(updated_items))

    @staticmethod
    def bulk_update_status(
        *,
        session: Session,
        payload: ScoreBulkStatusUpdatePayload,
    ) -> ScoreBulkStatusUpdateResponse:
        if not payload.scores:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="scores must not be empty.",
            )

        seen_ids: set[uuid.UUID] = set()
        for item in payload.scores:
            if item.id in seen_ids:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Duplicate score id in request payload.",
                )
            seen_ids.add(item.id)

        updated_items: list[Scores] = []
        for item in payload.scores:
            score = session.get(Scores, item.id)
            if not score:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Score not found with id={item.id}",
                )

            score.status = StatusEnum.ACTIVE
            score.updated_at = datetime.now()
            updated_items.append(score)

        session.commit()
        for item in updated_items:
            session.refresh(item)

        return ScoreBulkStatusUpdateResponse(items=updated_items, total=len(updated_items))

    @staticmethod
    def update(
        *, session: Session, score_id: uuid.UUID, score_data: ScoresUpdate
    ) -> ScoresPublic:
        score = session.get(Scores, score_id)
        if not score:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Score not found"
            )

        update_data = score_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(score, field, value)

        session.commit()
        session.refresh(score)
        if score.score_type is None:
            score.score_type = ScoreTypeEnum.OFFICIAL.value.capitalize()

        return ScoresPublic.model_validate(score)

    @staticmethod
    def delete(*, session: Session, score_id: uuid.UUID) -> ScoresDeleteResponse:
        score = session.get(Scores, score_id)
        if not score:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Score not found"
            )

        if score.status == StatusEnum.ACTIVE:
            score.status = StatusEnum.INACTIVE
            session.commit()
            return ScoresDeleteResponse(
                id=str(score.id), message="Score set to inactive"
            )

        session.delete(score)
        session.commit()

        return ScoresDeleteResponse(
            id=str(score.id), message="Score deleted successfully"
        )
