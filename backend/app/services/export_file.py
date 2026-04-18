from io import BytesIO
from typing import Any, Iterable, Sequence

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


class ExportFileServices:
    @staticmethod
    def export_excel_response(
        *,
        headers: Sequence[str],
        rows: Iterable[Sequence[Any]],
        file_name: str = "export.xlsx",
        sheet_name: str = "Sheet1",
    ) -> StreamingResponse:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # Write header row
        worksheet.append(list(headers))
        for column_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.font = Font(bold=True)

        # Write data rows
        for row in rows:
            worksheet.append(list(row))

        # Auto-fit column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        if not file_name.lower().endswith(".xlsx"):
            file_name = f"{file_name}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )

    @staticmethod
    def export_students_template_response(
        *,
        file_name: str = "students_template.xlsx",
    ) -> StreamingResponse:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Template"

        # Column widths to match template layout A..J
        column_widths = {
            "A": 6,   # TT
            "B": 20,  # Ma SV
            "C": 40,  # Ho va ten
            "D": 15,  # GT
            "E": 20,  # Lop
            "F": 20,  # Ngay sinh
            "G": 22,  # Noi sinh
            "H": 20,  # Dien thoai
            "I": 42,  # Email
            "J": 50,  # Ghi chu
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        # Header text blocks
        worksheet.merge_cells("B1:D1")
        worksheet["B1"] = "TRƯỜNG ĐHSPKT HƯNG YÊN"
        worksheet["B1"].font = Font(size=14, bold=True)
        worksheet["B1"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("G1:J1")
        worksheet["G1"] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
        worksheet["G1"].font = Font(size=14, bold=True)
        worksheet["G1"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("G2:J2")
        worksheet["G2"] = "Độc lập - Tự do - Hạnh phúc"
        worksheet["G2"].font = Font(size=14, bold=True, underline="single")
        worksheet["G2"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("B4:J4")
        worksheet["B4"] = "DANH SÁCH SINH VIÊN"
        worksheet["B4"].font = Font(size=22, bold=True)
        worksheet["B4"].alignment = Alignment(horizontal="center", vertical="center")

        # Table header row at row 6
        table_headers = ["TT", "Mã SV", "Họ và tên", "GT", "Lớp", "Ngày sinh", "Nơi sinh", "Điện thoại", "Email", "Ghi chú"]
        for col_index, header_value in enumerate(table_headers, start=1):
            cell = worksheet.cell(row=6, column=col_index, value=header_value)
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Template rows (blank student data, only serial number)
        template_rows = 30
        for row in range(7, 7 + template_rows):
            worksheet.cell(row=row, column=1, value=row - 6).alignment = Alignment(horizontal="center", vertical="center")
            for col in range(2, 11):
                worksheet.cell(row=row, column=col, value=None).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Borders around table area
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(6, 7 + template_rows):
            for col in range(1, 11):
                worksheet.cell(row=row, column=col).border = border

        # Row heights
        worksheet.row_dimensions[4].height = 44
        worksheet.row_dimensions[6].height = 32
        for row in range(7, 7 + template_rows):
            worksheet.row_dimensions[row].height = 28

        worksheet.freeze_panes = "A7"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        if not file_name.lower().endswith(".xlsx"):
            file_name = f"{file_name}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )

    @staticmethod
    def export_teachers_template_response(
        *,
        file_name: str = "teachers_template.xlsx",
    ) -> StreamingResponse:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Template"

        column_widths = {
            "A": 6,
            "B": 20,
            "C": 40,
            "D": 15,
            "E": 20,
            "F": 32,
            "G": 20,
            "H": 42,
            "I": 30,
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        worksheet.merge_cells("B1:D1")
        worksheet["B1"] = "TRƯỜNG ĐHSPKT HƯNG YÊN"
        worksheet["B1"].font = Font(size=14, bold=True)
        worksheet["B1"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("F1:I1")
        worksheet["F1"] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
        worksheet["F1"].font = Font(size=14, bold=True)
        worksheet["F1"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("F2:I2")
        worksheet["F2"] = "Độc lập - Tự do - Hạnh phúc"
        worksheet["F2"].font = Font(size=14, bold=True, underline="single")
        worksheet["F2"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("B4:I4")
        worksheet["B4"] = "DANH SÁCH GIẢNG VIÊN"
        worksheet["B4"].font = Font(size=22, bold=True)
        worksheet["B4"].alignment = Alignment(horizontal="center", vertical="center")

        table_headers = ["TT", "Mã GV", "Họ và tên", "GT", "Ngày sinh", "Nơi sinh", "Điện thoại", "Email", "Ghi chú"]
        for col_index, header_value in enumerate(table_headers, start=1):
            cell = worksheet.cell(row=6, column=col_index, value=header_value)
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        template_rows = 30
        for row in range(7, 7 + template_rows):
            worksheet.cell(row=row, column=1, value=row - 6).alignment = Alignment(horizontal="center", vertical="center")
            for col in range(2, 10):
                worksheet.cell(row=row, column=col, value=None).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(6, 7 + template_rows):
            for col in range(1, 10):
                worksheet.cell(row=row, column=col).border = border

        worksheet.row_dimensions[4].height = 44
        worksheet.row_dimensions[6].height = 32
        for row in range(7, 7 + template_rows):
            worksheet.row_dimensions[row].height = 28

        worksheet.freeze_panes = "A7"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        if not file_name.lower().endswith(".xlsx"):
            file_name = f"{file_name}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )

    @staticmethod
    def export_teaching_schedules_template_response(
        *,
        file_name: str = "teaching_schedules_template.xlsx",
    ) -> StreamingResponse:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Template"

        column_widths = {
            "A": 6,
            "B": 18,
            "C": 30,
            "D": 18,
            "E": 30,
            "F": 10,
            "G": 14,
            "H": 16,
            "I": 12,
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        worksheet.merge_cells("B1:E1")
        worksheet["B1"] = "PHIẾU NHẬP LỊCH DẠY"
        worksheet["B1"].font = Font(size=16, bold=True)
        worksheet["B1"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("B3:C3")
        worksheet["B3"] = "Mã lớp: CTK42A"
        worksheet["B3"].font = Font(bold=True)

        worksheet.merge_cells("B4:C4")
        worksheet["B4"] = "Giai đoạn học: dd/mm/yyyy - dd/mm/yyyy"
        worksheet["B4"].font = Font(bold=True)

        worksheet["B5"] = "Mã lớp"
        worksheet["B5"].font = Font(italic=True, color="666666")
        worksheet["C5"] = "CTK42A"
        worksheet["B7"] = "Niên khóa / học kỳ"
        worksheet["B7"].font = Font(italic=True, color="666666")
        worksheet["C7"] = "01/09/2025 - 31/12/2025"

        worksheet.merge_cells("A9:I9")
        worksheet["A9"] = "Dữ liệu bắt đầu từ dòng 10. Không sửa thứ tự cột."
        worksheet["A9"].font = Font(italic=True, color="666666")

        table_headers = [
            "TT",
            "Mã MH",
            "Tên MH",
            "Mã GV",
            "Tên GV",
            "Thứ",
            "Tiết Học",
            "Tuần học",
            "Phòng",
        ]
        for col_index, header_value in enumerate(table_headers, start=1):
            cell = worksheet.cell(row=10, column=col_index, value=header_value)
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        template_rows = 25
        for row in range(11, 11 + template_rows):
            worksheet.cell(row=row, column=1, value=row - 10).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            for col in range(2, 10):
                worksheet.cell(row=row, column=col, value=None).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(10, 11 + template_rows):
            for col in range(1, 10):
                worksheet.cell(row=row, column=col).border = border

        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[9].height = 22
        worksheet.row_dimensions[10].height = 32
        for row in range(11, 11 + template_rows):
            worksheet.row_dimensions[row].height = 26

        worksheet.freeze_panes = "A11"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        if not file_name.lower().endswith(".xlsx"):
            file_name = f"{file_name}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )
