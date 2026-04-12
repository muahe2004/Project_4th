import uuid
from datetime import datetime
from io import BytesIO

from app.models.schemas.common.query import BaseQueryParams, DateRange
from app.models.schemas.learning_schedules.learning_schedule_schemas import LearningSchedulePublic
from app.models.schemas.shared.teaching_schedule_embeds import TeachingScheduleClassInfo, TeachingScheduleInTeacher, TeachingScheduleRoomInfo, TeachingScheduleSubjectInfo
from fastapi import HTTPException, Request, UploadFile
from openpyxl import load_workbook
from sqlmodel import Session, and_, desc, func, or_, select
from starlette import status
from sqlalchemy import String, cast
from typing import List, Tuple
from app.middleware.hashing import hash_password

from app.models.models import Classes, LearningSchedules, Relatives, Rooms, Subjects, Teachers, TeachingSchedules, UserInformations
from app.models.schemas.relatives.relative_schemas import (
    RelativePublic,
    UserRelativeCreate,
)
from app.models.schemas.teachers.teacher_schemas import (
    TeacherCreateResponse,
    TeacherCreateWithUserInfor,
    TeacherDeleteResponse,
    TeacherDropdownResponse,
    TeacherFileData,
    TeacherFileDataResponse,
    TeacherFileInfo,
    TeacherFileInvalidRow,
    TeacherPublic,
    TeacherResponse,
    TeacherSearchParams,
    TeacherUploadField,
    TeacherUpdate,
    TeacherWithLearningSchedules,
)
from app.models.schemas.user_informations.user_information_schemas import (
    UserInformationPublic,
    UserInformationUpdate,
)
from app.services.departments import DepartmentServices
from app.services.common import build_date_conditions, parse_excel_datetime, to_clean_text
from app.services.user_information import User_Information_Services
from app.services.relatives import RelativeServices
from app.enums.status import StatusEnum


def _has_relative_payload_value(relative: UserRelativeCreate) -> bool:
    # `relatives.name` is required in DB, so only keep records with a non-empty name.
    return bool(relative.name and str(relative.name).strip())


def _sanitize_user_information_payload(payload: dict) -> dict:
    ignored_fields = {"student_id", "teacher_id", "created_at", "updated_at", "id"}
    return {key: value for key, value in payload.items() if key not in ignored_fields}


def _normalize_gender(raw_value: object) -> str | None:
    text_value = to_clean_text(raw_value)
    if not text_value:
        return None

    normalized = text_value.strip().lower()
    if normalized == "nam":
        return "1"
    if normalized in {"nữ", "nu"}:
        return "2"
    return "3"


def get_all_teachers() -> List[dict]:
    """Utility for services that need a lightweight teacher list (id, name)."""
    from app.core.database import engine

    with Session(engine) as session:
        teachers = session.exec(select(Teachers)).all()
        return [{"id": t.id, "name": t.name} for t in teachers]


class TeacherServices:
    @staticmethod
    def get_all(
        *,
        session: Session,
        query: TeacherSearchParams,
    ) -> tuple[list[TeacherResponse], int]:
        statement = select(Teachers)

        conditions = []
        if query.status:
            conditions.append(Teachers.status == query.status)

        if query.department_id:
            conditions.append(Teachers.department_id == query.department_id)

        if query.search:
            search_pattern = f"%{query.search}%"
            conditions.append(
                or_(
                    Teachers.teacher_code.ilike(search_pattern),
                    Teachers.name.ilike(search_pattern),
                )
            )

        if conditions:
            statement = statement.where(and_(*conditions))

        count_stmt = select(func.count()).select_from(Teachers)
        if conditions:
            count_stmt = count_stmt.where(and_(*conditions))
        total = session.exec(count_stmt).one()

        paged_statement = (
            statement.order_by(Teachers.created_at.desc())
            .offset(query.skip)
            .limit(query.limit)
        )
        teachers_page = session.exec(paged_statement).all()
        teacher_ids = [teacher.id for teacher in teachers_page]

        user_infos = {}
        if teacher_ids:
            infos = session.exec(
                select(UserInformations).where(
                    UserInformations.teacher_id.in_(teacher_ids)
                )
            ).all()
            user_infos = {info.teacher_id: info for info in infos}

        relatives_map = {}
        if teacher_ids:
            relatives = session.exec(
                select(Relatives).where(Relatives.teacher_id.in_(teacher_ids))
            ).all()
            for relative in relatives:
                relatives_map.setdefault(relative.teacher_id, []).append(relative)

        department_ids = [
            teacher.department_id
            for teacher in teachers_page
            if teacher.department_id is not None
        ]

        department_info = DepartmentServices.get_dropdown_by_ids(
            session=session,
            ids=department_ids,
            request=None,
        )
        department_map = {
            str(dep.id): {
                "department_code": dep.department_code,
                "department_name": dep.department_name,
            }
            for dep in department_info
        }

        teachers: list[TeacherResponse] = []
        for teacher in teachers_page:
            user_info = user_infos.get(teacher.id)
            teacher_relatives = relatives_map.get(teacher.id, [])
            dept_data = department_map.get(str(teacher.department_id), {})

            teachers.append(
                TeacherResponse(
                    **teacher.dict(),
                    department_code=dept_data.get("department_code"),
                    department_name=dept_data.get("department_name"),
                    teacher_information=(
                        UserInformationPublic.model_validate(user_info)
                        if user_info is not None
                        else None
                    ),
                    teacher_relative=[
                        RelativePublic.model_validate(rel) for rel in teacher_relatives
                    ]
                    if teacher_relatives
                    else None,
                )
            )

        return teachers, total

    # get teacher and learning schedule
    @staticmethod
    def get_teacher_and_learning_schedule(
        *, session: Session, query: BaseQueryParams, date_range: DateRange
    ) -> Tuple[List[TeacherWithLearningSchedules], int]:

        conditions = []
        schedule_conditions = build_date_conditions(date_range)

        if query.search:
            search_pattern = f"%{query.search}%"
            conditions.append(
                
                or_(
                    cast(Teachers.name, String).ilike(search_pattern),
                    cast(Teachers.teacher_code, String).ilike(search_pattern),
                )
            )

        count_stmt = select(func.count(Teachers.id))

        if conditions:
            count_stmt = count_stmt.where(and_(*conditions))

        total = session.exec(count_stmt).one()

        teacher_stmt = select(Teachers)

        if conditions:
            teacher_stmt = teacher_stmt.where(and_(*conditions))

        teacher_stmt = (
            teacher_stmt
            .order_by(desc(Teachers.created_at))
            .offset(query.skip)
            .limit(query.limit)
        )

        teachers = session.exec(teacher_stmt).all()

        if not teachers:
            return [], total

        teacher_ids = [r.id for r in teachers]

        stmt = (
            select(
                Teachers.id.label("teacher_id"),

                TeachingSchedules.id.label("ts_id"),
                TeachingSchedules.status.label("ts_status"),
                TeachingSchedules.created_at.label("ts_created_at"),
                TeachingSchedules.updated_at.label("ts_updated_at"),

                LearningSchedules.id.label("ls_id"),
                LearningSchedules.class_id.label("ls_class_id"),
                LearningSchedules.subject_id.label("ls_subject_id"),
                LearningSchedules.date.label("ls_date"),
                LearningSchedules.start_period.label("ls_start_period"),
                LearningSchedules.end_period.label("ls_end_period"),
                LearningSchedules.created_at.label("ls_created_at"),
                LearningSchedules.updated_at.label("ls_updated_at"),

                Rooms.id.label("room_id"),
                Rooms.room_number.label("room_number"),

                Classes.id.label("class_id"),
                Classes.class_name.label("class_name"),
                Classes.class_code.label("class_code"),

                Subjects.id.label("subject_id"),
                Subjects.name.label("subject_name"),
                Subjects.subject_code.label("subject_code"),
            )
            .select_from(Teachers)
            .outerjoin(
                TeachingSchedules,
                TeachingSchedules.teacher_id == Teachers.id
            )
            .outerjoin(
                LearningSchedules,
                LearningSchedules.id == TeachingSchedules.learning_schedule_id
            )
            .outerjoin(Rooms, Rooms.id == LearningSchedules.room_id)
            .outerjoin(Classes, Classes.id == LearningSchedules.class_id)
            .outerjoin(Subjects, Subjects.id == LearningSchedules.subject_id)
            .where(Teachers.id.in_(teacher_ids))
        )

        if schedule_conditions:
            stmt = stmt.where(and_(*schedule_conditions))

        rows = session.exec(stmt).all()

        teacher_map = {}

        for r in teachers:
            teacher_map[r.id] = TeacherWithLearningSchedules(
                teacher_information=TeacherPublic.model_validate(r),
                teaching_schedules=[]
            )

        for row in rows:

            if row.ls_id is None:
                continue

            schedule_item = TeachingScheduleInTeacher(
                id=row.ts_id or row.ls_id,
                status=row.ts_status,  
                created_at=row.ts_created_at or row.ls_created_at,
                updated_at=row.ts_updated_at or row.ls_updated_at,

                learning_schedule=LearningSchedulePublic(
                    id=row.ls_id,
                    class_id=row.ls_class_id,
                    subject_id=row.ls_subject_id,
                    date=row.ls_date,
                    start_period=row.ls_start_period,
                    end_period=row.ls_end_period,
                    created_at=row.ls_created_at,
                    updated_at=row.ls_updated_at,
                ),

                room = (
                    TeachingScheduleRoomInfo(
                        room_id=row.room_id,
                        room_number=row.room_number,
                    )
                    if row.room_id else None
                ),

                class_info=(
                    TeachingScheduleClassInfo(
                        class_id=row.class_id,
                        class_name=row.class_name,
                        class_code=row.class_code,
                    )
                    if row.class_id else None
                ),

                subject=(
                    TeachingScheduleSubjectInfo(
                        subject_id=row.subject_id,
                        subject_name=row.subject_name,
                        subject_code=row.subject_code,
                    )
                    if row.subject_id else None
                ),
            )

            teacher_map[row.teacher_id].teaching_schedules.append(schedule_item)

        return list(teacher_map.values()), total
    
    @staticmethod
    def get_list_teacher(
        *, session: Session, teacher_ids: List[str]
    ) -> List[TeacherResponse]:
        if not teacher_ids:
            return []

        statement = select(Teachers).where(Teachers.id.in_(teacher_ids))
        results = session.exec(statement).all()
        return results

    @staticmethod
    def get_dropdown(
        *, session: Session, query: TeacherSearchParams
    ) -> List[TeacherDropdownResponse]:
        statement = select(Teachers)

        conditions = []
        if query.status:
            conditions.append(Teachers.status == query.status)

        if query.department_id:
            conditions.append(Teachers.department_id == query.department_id)

        if query.search:
            search_pattern = f"%{query.search}%"
            conditions.append(
                or_(
                    Teachers.teacher_code.ilike(search_pattern),
                    Teachers.name.ilike(search_pattern),
                )
            )

        if conditions:
            statement = statement.where(and_(*conditions))

        statement = statement.order_by(desc(Teachers.created_at))
        statement = statement.offset(query.skip).limit(query.limit)

        results = session.exec(statement).all()
        return [TeacherDropdownResponse.model_validate(t) for t in results]

    @staticmethod
    def get_dropdown_by_ids(
        *, session: Session, ids: List[uuid.UUID], request: Request
    ) -> List[TeacherDropdownResponse]:
        if not ids:
            return []

        statement = select(Teachers).where(Teachers.id.in_(ids))
        results = session.exec(statement).all()
        return [TeacherDropdownResponse.model_validate(t) for t in results]

    @staticmethod
    def get_by_id(
        *, session: Session, teacher_id: uuid.UUID, request: Request
    ) -> TeacherPublic:
        teacher = session.get(Teachers, teacher_id)
        if not teacher:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Teacher does not exist"
            )
        return TeacherPublic.model_validate(teacher)

    @staticmethod
    def create(
        *,
        session: Session,
        teacher: TeacherCreateWithUserInfor,
    ) -> TeacherCreateResponse:
        existing_teacher_code = session.exec(
            select(Teachers).where(Teachers.teacher_code == teacher.teacher_code)
        ).first()
        if existing_teacher_code:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Teacher already exists",
            )
        existing_email = session.exec(
            select(Teachers).where(Teachers.email == teacher.email)
        ).first()
        if existing_email:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Teacher email already exists",
            )

        teacher_data = teacher.model_dump(
            exclude={"teacher_information", "teacher_relatives"}
        )
        teacher_data["password"] = hash_password(teacher_data["password"])

        new_teacher = Teachers(**teacher_data)
        session.add(new_teacher)
        session.flush()

        user_info_payload = _sanitize_user_information_payload(
            teacher.teacher_information.model_dump(exclude_none=True)
        )
        user_info = UserInformations(**user_info_payload, teacher_id=new_teacher.id)
        session.add(user_info)

        relatives = []
        for relative in teacher.teacher_relatives:
            relative_data = relative.model_dump(exclude_none=True)
            if not relative_data:
                continue
            relative_record = Relatives(
                **relative_data,
                teacher_id=new_teacher.id,
            )
            session.add(relative_record)
            relatives.append(relative_record)

        session.commit()
        session.refresh(new_teacher)
        session.refresh(user_info)

        teacher_relatives_response = (
            [RelativePublic.model_validate(rel) for rel in relatives]
            if relatives
            else None
        )

        return TeacherCreateResponse(
            **new_teacher.dict(),
            teacher_information=UserInformationPublic.model_validate(user_info),
            teacher_relative=teacher_relatives_response,
        )

    @staticmethod
    async def upload_file_teacher(
        *,
        session: Session,
        file: UploadFile,
    ) -> TeacherFileDataResponse:
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel sheet is empty.",
            )

        expected_header_labels = {
            TeacherUploadField.CODE: "Mã GV",
            TeacherUploadField.NAME: "Họ và tên",
            TeacherUploadField.GENDER: "GT",
            TeacherUploadField.DATE_OF_BIRTH: "Ngày sinh",
            TeacherUploadField.ADDRESS: "Nơi sinh",
            TeacherUploadField.PHONE: "Điện thoại",
            TeacherUploadField.EMAIL: "Email",
        }
        header_row_index = 6
        if len(rows) < header_row_index:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"File does not contain header row {header_row_index}. "
                    "Please use the correct template."
                ),
            )

        fixed_column_indexes = {
            TeacherUploadField.CODE: 1,
            TeacherUploadField.NAME: 2,
            TeacherUploadField.GENDER: 3,
            TeacherUploadField.DATE_OF_BIRTH: 4,
            TeacherUploadField.ADDRESS: 5,
            TeacherUploadField.PHONE: 6,
            TeacherUploadField.EMAIL: 7,
        }

        parsed_rows: list[TeacherFileData] = []
        invalid_rows: list[TeacherFileInvalidRow] = []

        for row_index, row_values in enumerate(rows[header_row_index:], start=header_row_index + 1):
            if row_values is None:
                continue

            def cell(field_name: TeacherUploadField) -> object | None:
                col_index = fixed_column_indexes[field_name]
                if col_index >= len(row_values):
                    return None
                return row_values[col_index]

            teacher_code = to_clean_text(cell(TeacherUploadField.CODE))
            teacher_name = to_clean_text(cell(TeacherUploadField.NAME))
            gender = _normalize_gender(cell(TeacherUploadField.GENDER))
            date_of_birth_raw = cell(TeacherUploadField.DATE_OF_BIRTH)
            address = to_clean_text(cell(TeacherUploadField.ADDRESS))
            phone = to_clean_text(cell(TeacherUploadField.PHONE))
            email = to_clean_text(cell(TeacherUploadField.EMAIL))

            if all(
                value in (None, "")
                for value in (
                    teacher_code,
                    teacher_name,
                    date_of_birth_raw,
                    address,
                    phone,
                    email,
                )
            ):
                continue

            try:
                date_of_birth = parse_excel_datetime(date_of_birth_raw)
            except ValueError as exc:
                invalid_rows.append(
                    TeacherFileInvalidRow(
                        row=row_index,
                        teacher_code=teacher_code,
                        name=teacher_name,
                        gender=gender,
                        date_of_birth=None,
                        email=email,
                        phone=phone,
                        address=address,
                        errors=[str(exc)],
                    )
                )
                continue

            row_errors: list[str] = []
            if not teacher_code:
                row_errors.append("Teacher Code is required.")
            if not teacher_name:
                row_errors.append("Teacher Name is required.")
            if not gender:
                row_errors.append("Gender is required.")
            if not email:
                row_errors.append("Email is required.")

            if row_errors:
                invalid_rows.append(
                    TeacherFileInvalidRow(
                        row=row_index,
                        teacher_code=teacher_code,
                        name=teacher_name,
                        gender=gender,
                        date_of_birth=date_of_birth,
                        email=email,
                        phone=phone,
                        address=address,
                        errors=row_errors,
                    )
                )
                continue

            parsed_rows.append(
                TeacherFileData(
                    teacher_code=teacher_code,
                    name=teacher_name,
                    gender=gender,
                    date_of_birth=date_of_birth,
                    email=email,
                    phone=phone,
                    address=address,
                )
            )

        return TeacherFileDataResponse(
            file_information=TeacherFileInfo(
                file_name=filename,
                headers=list(expected_header_labels.values()),
                header_row=header_row_index,
                total_rows=max(len(rows) - header_row_index, 0),
                valid_rows_count=len(parsed_rows),
                invalid_rows_count=len(invalid_rows),
            ),
            teachers=parsed_rows,
            invalid_teachers=invalid_rows,
        )

    @staticmethod
    def import_list_teacher(
        *,
        session: Session,
        list_teacher: List[TeacherFileData],
    ) -> List[TeacherFileData]:
        imported_teachers: list[TeacherFileData] = []

        for teacher_item in list_teacher:
            teacher_code = (teacher_item.teacher_code or "").strip()
            teacher_name = (teacher_item.name or "").strip()
            email = (teacher_item.email or "").strip()
            gender = (teacher_item.gender or "").strip()

            if not teacher_code or not teacher_name or not email or not gender:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="teacher_code, name, gender, email are required.",
                )
            if gender not in {"1", "2", "3"}:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="gender must be one of: 1, 2, 3.",
                )

            existed_teacher = session.exec(
                select(Teachers).where(
                    or_(
                        Teachers.teacher_code == teacher_code,
                        Teachers.email == email,
                    )
                )
            ).first()
            if existed_teacher:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Teacher already exists: {teacher_code} / {email}",
                )

            new_teacher = Teachers(
                teacher_code=teacher_code,
                name=teacher_name,
                date_of_birth=teacher_item.date_of_birth,
                gender=gender,
                email=email,
                phone=teacher_item.phone,
                address=teacher_item.address,
                academic_rank=None,
                status=StatusEnum.ACTIVE,
                department_id=None,
                password=hash_password(teacher_code),
                created_at=datetime.now(),
                updated_at=datetime.now(),
            )
            session.add(new_teacher)

            imported_teachers.append(
                TeacherFileData(
                    teacher_code=new_teacher.teacher_code,
                    name=new_teacher.name,
                    gender=new_teacher.gender,
                    date_of_birth=new_teacher.date_of_birth,
                    email=new_teacher.email,
                    phone=new_teacher.phone,
                    address=new_teacher.address,
                )
            )

        session.commit()
        return imported_teachers

    @staticmethod
    def update(
        *,
        session: Session,
        teacher_id: uuid.UUID,
        teacher_data: TeacherUpdate,
    ) -> TeacherPublic:
        teacher = session.get(Teachers, teacher_id)
        if not teacher:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Teacher not found"
            )

        if (
            teacher_data.teacher_code
            and teacher_data.teacher_code != teacher.teacher_code
            and session.exec(
                select(Teachers).where(Teachers.teacher_code == teacher_data.teacher_code)
            ).first()
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Teacher code already exists",
            )

        if (
            teacher_data.email
            and teacher_data.email != teacher.email
            and session.exec(
                select(Teachers).where(
                    Teachers.email == teacher_data.email, Teachers.id != teacher_id
                )
            ).first()
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Teacher email already exists",
            )

        update_data = teacher_data.model_dump(
            exclude_unset=True,
            exclude={"teacher_information", "teacher_relatives"},
        )
        for field, value in update_data.items():
            setattr(teacher, field, value)
        if teacher_data.teacher_information:
            info_payload = _sanitize_user_information_payload(
                teacher_data.teacher_information.model_dump(exclude_none=True)
            )
            if info_payload:
                user_info = session.exec(
                    select(UserInformations).where(
                        UserInformations.teacher_id == teacher.id
                    )
                ).one_or_none()
                user_info_update = UserInformationUpdate(**info_payload)
                if user_info:
                    User_Information_Services.update(
                        session=session,
                        user_information_id=user_info.id,
                        user_information_data=user_info_update,
                        commit=False,
                    )
                else:
                    session.add(UserInformations(**info_payload, teacher_id=teacher.id))

        if teacher_data.teacher_relatives is not None:
            filtered_relatives = [
                rel
                for rel in teacher_data.teacher_relatives
                if _has_relative_payload_value(rel)
            ]
            RelativeServices.replace_for_teacher(
                session=session,
                teacher_id=teacher.id,
                relatives=filtered_relatives,
                commit=False,
            )

        session.commit()
        return TeacherPublic.model_validate(teacher)

    @staticmethod
    def delete_many(
        *, session: Session, teacher_ids: List[uuid.UUID]
    ) -> List[TeacherDeleteResponse]:
        results: List[TeacherDeleteResponse] = []

        for teacher_id in teacher_ids:
            teacher = session.get(Teachers, teacher_id)
            if not teacher:
                results.append(
                    TeacherDeleteResponse(
                        id=str(teacher_id), message="Teacher not found"
                    )
                )
                continue

            check_related_entities = select(Classes).where(
                Classes.teacher_id == teacher.id
            )
            classes = session.exec(check_related_entities).all()
            if classes:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Teacher has related classes and cannot be deleted.",
                )

            if teacher.status != StatusEnum.INACTIVE:
                teacher.status = StatusEnum.INACTIVE
                message = "Teacher set to inactive"
            else:
                message = "Teacher already inactive"

            session.commit()
            results.append(TeacherDeleteResponse(id=str(teacher_id), message=message))

        return results
