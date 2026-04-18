import uuid
from io import BytesIO
import re
from datetime import datetime

from fastapi import HTTPException, Request, UploadFile
from sqlmodel import Session, select
from starlette import status
from typing import List
from sqlalchemy.orm import aliased
from sqlalchemy import String, cast, func, or_
from openpyxl import load_workbook

from app.enums.status import StatusEnum
from app.models.models import (
    ExaminationSchedules,
    Classes,
    Subjects,
    Rooms,
    Teachers,
    StudentClass,
)
from app.models.schemas.examination_schedules.examination_schedule_schemas import (
    ExaminationSchedulePublic,
    ExaminationScheduleCreate,
    ExaminationScheduleUpdate,
    ExaminationScheduleDeleteResponse,
    ExaminationScheduleResponse,
    ExaminationScheduleClassInfo,
    ExaminationScheduleSubjectInfo,
    ExaminationScheduleRoomInfo,
    ExaminationScheduleInvigilatorInfo,
    ExaminationScheduleQueryParams,
    UploadExaminationScheduleResponse,
    UploadExaminationScheduleFileInfo,
    UploadExaminationScheduleItem,
    UploadExaminationScheduleInvalidRow,
    ImportExaminationScheduleInput,
    ImportExaminationScheduleResponse,
    ImportExaminationScheduleImportedItem,
)
from app.models.schemas.common.query import DateRange
from app.services.common import build_date_conditions_for_column
from app.services.common import parse_excel_datetime, to_clean_text


class ExaminationScheduleServices:
    @staticmethod
    async def upload_file_examination_schedule(
        *,
        session: Session,
        file: UploadFile,
    ) -> UploadExaminationScheduleResponse:
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        workbook = load_workbook(BytesIO(content), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < 6:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not contain the expected header row.",
            )

        expected_headers = [
            "TT",
            "Mã Môn Thi",
            "Tên Môn Thi",
            "Mã Lớp",
            "Mã Giám Thi 1",
            "Tên Giám Thi 1",
            "Mã Giám Thi 2",
            "Tên Giám Thi 2",
            "Phòng",
            "Ngày Thi",
            "Giờ Bắt Đầu",
            "Giờ Kết Thúc",
            "Loại",
        ]

        header_row_index = 5
        parsed_rows: list[UploadExaminationScheduleItem] = []
        invalid_rows: list[UploadExaminationScheduleInvalidRow] = []

        subject_codes = {
            to_clean_text(row[1])
            for row in rows[header_row_index + 1:]
            if row and len(row) > 1 and to_clean_text(row[1])
        }
        class_codes = {
            to_clean_text(row[3])
            for row in rows[header_row_index + 1:]
            if row and len(row) > 3 and to_clean_text(row[3])
        }
        teacher_codes = {
            code
            for row in rows[header_row_index + 1:]
            for code in (
                to_clean_text(row[4]) if row and len(row) > 4 else None,
                to_clean_text(row[6]) if row and len(row) > 6 else None,
            )
            if code
        }
        room_numbers = {
            int(room_value)
            for row in rows[header_row_index + 1:]
            if row and len(row) > 8 and (room_value := to_clean_text(row[8])) and str(room_value).isdigit()
        }

        subject_records = session.exec(
            select(Subjects).where(Subjects.subject_code.in_(list(subject_codes)))
        ).all() if subject_codes else []
        class_records = session.exec(
            select(Classes).where(Classes.class_code.in_(list(class_codes)))
        ).all() if class_codes else []
        teacher_records = session.exec(
            select(Teachers).where(Teachers.teacher_code.in_(list(teacher_codes)))
        ).all() if teacher_codes else []
        room_records = session.exec(
            select(Rooms).where(Rooms.room_number.in_(list(room_numbers)))
        ).all() if room_numbers else []

        subject_map = {item.subject_code: item for item in subject_records}
        class_map = {item.class_code: item for item in class_records}
        teacher_map = {item.teacher_code: item for item in teacher_records}
        room_map = {item.room_number: item for item in room_records}

        for row_index, row_values in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            if row_values is None:
                continue

            def cell(col_index: int) -> object | None:
                if col_index >= len(row_values):
                    return None
                return row_values[col_index]

            subject_code = to_clean_text(cell(1))
            subject_name = to_clean_text(cell(2))
            class_code = to_clean_text(cell(3))
            invigilator_1_code = to_clean_text(cell(4))
            invigilator_1_name = to_clean_text(cell(5))
            invigilator_2_code = to_clean_text(cell(6))
            invigilator_2_name = to_clean_text(cell(7))
            room_raw = to_clean_text(cell(8))
            date_raw = cell(9)
            start_time_raw = cell(10)
            end_time_raw = cell(11)
            schedule_type = to_clean_text(cell(12))

            if all(
                value in (None, "")
                for value in (
                    subject_code,
                    subject_name,
                    class_code,
                    invigilator_1_code,
                    invigilator_1_name,
                    invigilator_2_code,
                    invigilator_2_name,
                    room_raw,
                    date_raw,
                    start_time_raw,
                    end_time_raw,
                    schedule_type,
                )
            ):
                continue

            row_errors: list[str] = []
            room_number: int | None = None
            date_value: datetime | None = None
            start_time_value: datetime | None = None
            end_time_value: datetime | None = None

            if not subject_code:
                row_errors.append("Subject Code is required.")
            if not class_code:
                row_errors.append("Class Code is required.")
            if not room_raw:
                row_errors.append("Room is required.")
            if not date_raw:
                row_errors.append("Exam date is required.")
            if not start_time_raw:
                row_errors.append("Start time is required.")
            if not end_time_raw:
                row_errors.append("End time is required.")

            if room_raw:
                if str(room_raw).isdigit():
                    room_number = int(room_raw)
                else:
                    row_errors.append("Room must be a number.")

            try:
                date_value = parse_excel_datetime(date_raw)
            except ValueError:
                row_errors.append("Invalid exam date format.")

            try:
                start_time_value = parse_excel_datetime(start_time_raw)
            except ValueError:
                row_errors.append("Invalid start time format.")

            try:
                end_time_value = parse_excel_datetime(end_time_raw)
            except ValueError:
                row_errors.append("Invalid end time format.")

            if date_value and start_time_value and end_time_value and end_time_value <= start_time_value:
                row_errors.append("End time must be greater than start time.")

            subject_record = subject_map.get(subject_code) if subject_code else None
            class_record = class_map.get(class_code) if class_code else None
            invigilator_1_record = teacher_map.get(invigilator_1_code) if invigilator_1_code else None
            invigilator_2_record = teacher_map.get(invigilator_2_code) if invigilator_2_code else None
            room_record = room_map.get(room_number) if room_number is not None else None

            if subject_code and subject_record is None:
                row_errors.append(f"Subject not found with subject_code={subject_code}")
            if class_code and class_record is None:
                row_errors.append(f"Class not found with class_code={class_code}")
            if invigilator_1_code and invigilator_1_record is None:
                row_errors.append(f"Teacher not found with teacher_code={invigilator_1_code}")
            if invigilator_2_code and invigilator_2_record is None:
                row_errors.append(f"Teacher not found with teacher_code={invigilator_2_code}")
            if room_number is not None and room_record is None:
                row_errors.append(f"Room not found with room_number={room_number}")

            item = UploadExaminationScheduleItem(
                subject_id=subject_record.id if subject_record else None,
                subject_code=subject_code,
                subject_name=subject_name or (subject_record.name if subject_record else None),
                class_id=class_record.id if class_record else None,
                class_code=class_code,
                class_name=class_record.class_name if class_record else None,
                invigilator_1_id=invigilator_1_record.id if invigilator_1_record else None,
                invigilator_1_code=invigilator_1_code,
                invigilator_1_name=invigilator_1_name or (invigilator_1_record.name if invigilator_1_record else None),
                invigilator_2_id=invigilator_2_record.id if invigilator_2_record else None,
                invigilator_2_code=invigilator_2_code,
                invigilator_2_name=invigilator_2_name or (invigilator_2_record.name if invigilator_2_record else None),
                room_id=room_record.id if room_record else None,
                room_number=room_number,
                date=date_value,
                start_time=start_time_value,
                end_time=end_time_value,
                schedule_type=schedule_type,
            )

            if row_errors:
                invalid_rows.append(
                    UploadExaminationScheduleInvalidRow(
                        **item.model_dump(),
                        row=row_index,
                        errors=row_errors,
                    )
                )
                continue

            parsed_rows.append(item)

        return UploadExaminationScheduleResponse(
            file_information=UploadExaminationScheduleFileInfo(
                file_name=filename,
                headers=expected_headers,
                header_row=header_row_index,
                total_rows=max(len(rows) - (header_row_index + 1), 0),
                valid_rows_count=len(parsed_rows),
                invalid_rows_count=len(invalid_rows),
            ),
            schedules=parsed_rows,
            invalid_schedules=invalid_rows,
        )

    @staticmethod
    def import_examination_schedule(
        *,
        session: Session,
        calender: ImportExaminationScheduleInput,
    ) -> ImportExaminationScheduleResponse:
        imported_items: list[ImportExaminationScheduleImportedItem] = []
        pending_schedules: list[dict] = []

        try:
            for row_index, schedule in enumerate(calender.schedules, start=1):
                if not schedule.subject_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing subject_id at row {row_index}. Please upload/resolve file before import.",
                    )
                if not schedule.class_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing class_id at row {row_index}. Please upload/resolve file before import.",
                    )
                if not schedule.date or not schedule.start_time or not schedule.end_time:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing date/time at row {row_index}. Please upload/resolve file before import.",
                    )

                if schedule.end_time <= schedule.start_time:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"End time must be greater than start time at row {row_index}.",
                    )

                pending_schedules.append(
                    {
                        "row": row_index,
                        "class_id": schedule.class_id,
                        "subject_id": schedule.subject_id,
                        "date": schedule.date,
                        "start_time": schedule.start_time,
                        "end_time": schedule.end_time,
                        "room_id": schedule.room_id,
                        "schedule_type": schedule.schedule_type,
                        "status": schedule.status,
                        "invigilator_1_id": schedule.invigilator_1_id,
                        "invigilator_2_id": schedule.invigilator_2_id,
                    }
                )

            if pending_schedules:
                class_ids = list({item["class_id"] for item in pending_schedules})
                subject_ids = list({item["subject_id"] for item in pending_schedules})
                dates = [item["date"] for item in pending_schedules]
                min_date = min(dates)
                max_date = max(dates)

                existing_schedules = session.exec(
                    select(ExaminationSchedules).where(
                        ExaminationSchedules.class_id.in_(class_ids),
                        ExaminationSchedules.subject_id.in_(subject_ids),
                        ExaminationSchedules.date >= min_date,
                        ExaminationSchedules.date <= max_date,
                        or_(
                            ExaminationSchedules.status == StatusEnum.ACTIVE,
                            ExaminationSchedules.status.is_(None),
                        ),
                    )
                ).all()

                existing_group_map: dict[tuple[uuid.UUID, uuid.UUID, datetime], list[ExaminationSchedules]] = {}
                for existing_item in existing_schedules:
                    group_key = (existing_item.class_id, existing_item.subject_id, existing_item.date)
                    existing_group_map.setdefault(group_key, []).append(existing_item)

                for pending_item in pending_schedules:
                    class_record = session.get(Classes, pending_item["class_id"])
                    subject_record = session.get(Subjects, pending_item["subject_id"])
                    if not class_record:
                        raise HTTPException(
                            status_code=status.HTTP_404_NOT_FOUND,
                            detail=f"Class not found with class_id={pending_item['class_id']}",
                        )
                    if not subject_record:
                        raise HTTPException(
                            status_code=status.HTTP_404_NOT_FOUND,
                            detail=f"Subject not found with subject_id={pending_item['subject_id']}",
                        )

                    group_key = (pending_item["class_id"], pending_item["subject_id"], pending_item["date"])
                    for existing_item in existing_group_map.get(group_key, []):
                        if (
                            existing_item.start_time <= pending_item["end_time"]
                            and existing_item.end_time >= pending_item["start_time"]
                        ):
                            raise HTTPException(
                                status_code=status.HTTP_400_BAD_REQUEST,
                                detail=(
                                    f"Examination schedule overlaps with existing schedule for class "
                                    f"{class_record.class_code} and subject {subject_record.subject_code} "
                                    f"on {pending_item['date'].strftime('%d/%m/%Y')}."
                                ),
                            )

            for item in pending_schedules:
                examination_schedule = ExaminationSchedules(
                    class_id=item["class_id"],
                    subject_id=item["subject_id"],
                    date=item["date"],
                    start_time=item["start_time"],
                    end_time=item["end_time"],
                    room_id=item["room_id"],
                    schedule_type=item["schedule_type"],
                    status=item["status"] or StatusEnum.ACTIVE,
                    invigilator_1_id=item["invigilator_1_id"],
                    invigilator_2_id=item["invigilator_2_id"],
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                )
                session.add(examination_schedule)
                session.flush()

                imported_items.append(
                    ImportExaminationScheduleImportedItem(
                        row=item["row"],
                        id=examination_schedule.id,
                    )
                )

            session.commit()
        except HTTPException:
            session.rollback()
            raise
        except Exception:
            session.rollback()
            raise

        return ImportExaminationScheduleResponse(items=imported_items)

    @staticmethod
    def _build_response(
        *,
        row: dict,
    ) -> ExaminationScheduleResponse:
        invigilator_map: dict[str, ExaminationScheduleInvigilatorInfo] = {}

        for invigilator_id_key, code_key, name_key, email_key, phone_key in (
            (
                "invigilator_1_id",
                "invigilator_1_code",
                "invigilator_1_name",
                "invigilator_1_email",
                "invigilator_1_phone_number",
            ),
            (
                "invigilator_2_id",
                "invigilator_2_code",
                "invigilator_2_name",
                "invigilator_2_email",
                "invigilator_2_phone_number",
            ),
        ):
            invigilator_id = row.get(invigilator_id_key)
            if not invigilator_id:
                continue

            invigilator_map[str(invigilator_id)] = ExaminationScheduleInvigilatorInfo(
                invigilator_id=invigilator_id,
                invigilator_code=row.get(code_key),
                invigilator_name=row.get(name_key),
                invigilator_email=row.get(email_key),
                invigilator_phone_number=row.get(phone_key),
            )

        class_info = (
            ExaminationScheduleClassInfo(
                class_id=row["class_id"],
                class_code=row.get("class_code"),
                class_name=row.get("class_name"),
            )
            if row.get("class_id")
            else None
        )
        subject_info = (
            ExaminationScheduleSubjectInfo(
                subject_id=row["subject_id"],
                subject_code=row.get("subject_code"),
                subject_name=row.get("subject_name"),
            )
            if row.get("subject_id")
            else None
        )
        room_info = (
            ExaminationScheduleRoomInfo(
                room_id=row["room_id"],
                room_number=row.get("room_number"),
            )
            if row.get("room_id")
            else None
        )

        return ExaminationScheduleResponse(
            id=row["id"],
            date=row["date"],
            start_time=row["start_time"],
            end_time=row["end_time"],
            status=row.get("status"),
            schedule_type=row.get("schedule_type"),
            class_info=class_info,
            subject_info=subject_info,
            room_info=room_info,
            invigilator=list(invigilator_map.values()),
        )

    @staticmethod
    def _build_query_statement(
        *, query: ExaminationScheduleQueryParams, date_range: DateRange
    ):
        teacher_one = aliased(Teachers)
        teacher_two = aliased(Teachers)
        student_class_subquery = None

        if query.student_id:
            student_class_subquery = (
                select(StudentClass.class_id)
                .where(
                    StudentClass.student_id == query.student_id,
                    (
                        (StudentClass.status == StatusEnum.ACTIVE)
                        | (StudentClass.status.is_(None))
                    ),
                )
                .distinct()
                .subquery()
            )

        statement = (
            select(
                ExaminationSchedules.id.label("id"),
                ExaminationSchedules.date.label("date"),
                ExaminationSchedules.start_time.label("start_time"),
                ExaminationSchedules.end_time.label("end_time"),
                ExaminationSchedules.status.label("status"),
                ExaminationSchedules.schedule_type.label("schedule_type"),
                Classes.id.label("class_id"),
                Classes.class_code.label("class_code"),
                Classes.class_name.label("class_name"),
                Subjects.id.label("subject_id"),
                Subjects.subject_code.label("subject_code"),
                Subjects.name.label("subject_name"),
                Rooms.id.label("room_id"),
                Rooms.room_number.label("room_number"),
                teacher_one.id.label("invigilator_1_id"),
                teacher_one.teacher_code.label("invigilator_1_code"),
                teacher_one.name.label("invigilator_1_name"),
                teacher_one.email.label("invigilator_1_email"),
                teacher_one.phone.label("invigilator_1_phone_number"),
                teacher_two.id.label("invigilator_2_id"),
                teacher_two.teacher_code.label("invigilator_2_code"),
                teacher_two.name.label("invigilator_2_name"),
                teacher_two.email.label("invigilator_2_email"),
                teacher_two.phone.label("invigilator_2_phone_number"),
            )
            .select_from(ExaminationSchedules)
            .join(Classes, Classes.id == ExaminationSchedules.class_id)
            .join(Subjects, Subjects.id == ExaminationSchedules.subject_id)
            .outerjoin(Rooms, Rooms.id == ExaminationSchedules.room_id)
            .outerjoin(teacher_one, teacher_one.id == ExaminationSchedules.invigilator_1_id)
            .outerjoin(teacher_two, teacher_two.id == ExaminationSchedules.invigilator_2_id)
        )

        if student_class_subquery is not None:
            statement = statement.join(
                student_class_subquery,
                student_class_subquery.c.class_id == ExaminationSchedules.class_id,
            )

        conditions = []

        if query.status:
            conditions.append(ExaminationSchedules.status == query.status)

        if query.subject_id:
            conditions.append(ExaminationSchedules.subject_id == query.subject_id)

        if query.class_id:
            conditions.append(ExaminationSchedules.class_id == query.class_id)

        if query.invigilator_id:
            conditions.append(
                (
                    ExaminationSchedules.invigilator_1_id == query.invigilator_id
                )
                | (
                    ExaminationSchedules.invigilator_2_id == query.invigilator_id
                )
            )

        conditions.extend(
            build_date_conditions_for_column(date_range, ExaminationSchedules.date)
        )

        if query.search:
            search_pattern = f"%{query.search}%"
            conditions.append(
                (
                    cast(Rooms.room_number, String).ilike(search_pattern)
                    | Subjects.subject_code.ilike(search_pattern)
                    | Subjects.name.ilike(search_pattern)
                )
            )

        if conditions:
            statement = statement.where(*conditions)

        return statement, conditions

    @staticmethod
    def get_all(
        *, session: Session, query: ExaminationScheduleQueryParams, date_range: DateRange
    ) -> tuple[list[ExaminationScheduleResponse], int]:
        statement, conditions = ExaminationScheduleServices._build_query_statement(
            query=query, date_range=date_range
        )
        student_class_subquery = None
        if query.student_id:
            student_class_subquery = (
                select(StudentClass.class_id)
                .where(
                    StudentClass.student_id == query.student_id,
                    (
                        (StudentClass.status == StatusEnum.ACTIVE)
                        | (StudentClass.status.is_(None))
                    ),
                )
                .distinct()
                .subquery()
            )

        count_stmt = (
            select(func.count())
            .select_from(ExaminationSchedules)
            .join(Classes, Classes.id == ExaminationSchedules.class_id)
            .join(Subjects, Subjects.id == ExaminationSchedules.subject_id)
            .outerjoin(Rooms, Rooms.id == ExaminationSchedules.room_id)
        )
        if student_class_subquery is not None:
            count_stmt = count_stmt.join(
                student_class_subquery,
                student_class_subquery.c.class_id == ExaminationSchedules.class_id,
            )
        if conditions:
            count_stmt = count_stmt.where(*conditions)
        total = session.exec(count_stmt).one()

        statement = statement.order_by(
            ExaminationSchedules.date.desc(),
            ExaminationSchedules.created_at.desc(),
        ).offset(query.skip).limit(query.limit)

        rows = session.exec(statement).all()
        examination_schedules = [
            ExaminationScheduleServices._build_response(row=row._asdict())
            for row in rows
        ]
        return examination_schedules, total

    @staticmethod
    def get_by_id(
        *, session: Session, examination_schedules_id: uuid.UUID, request: Request
    ) -> ExaminationScheduleResponse:
        teacher_one = aliased(Teachers)
        teacher_two = aliased(Teachers)
        row = session.exec(
            select(
                ExaminationSchedules.id.label("id"),
                ExaminationSchedules.date.label("date"),
                ExaminationSchedules.start_time.label("start_time"),
                ExaminationSchedules.end_time.label("end_time"),
                ExaminationSchedules.status.label("status"),
                ExaminationSchedules.schedule_type.label("schedule_type"),
                Classes.id.label("class_id"),
                Classes.class_code.label("class_code"),
                Classes.class_name.label("class_name"),
                Subjects.id.label("subject_id"),
                Subjects.subject_code.label("subject_code"),
                Subjects.name.label("subject_name"),
                Rooms.id.label("room_id"),
                Rooms.room_number.label("room_number"),
                teacher_one.id.label("invigilator_1_id"),
                teacher_one.teacher_code.label("invigilator_1_code"),
                teacher_one.name.label("invigilator_1_name"),
                teacher_one.email.label("invigilator_1_email"),
                teacher_one.phone.label("invigilator_1_phone_number"),
                teacher_two.id.label("invigilator_2_id"),
                teacher_two.teacher_code.label("invigilator_2_code"),
                teacher_two.name.label("invigilator_2_name"),
                teacher_two.email.label("invigilator_2_email"),
                teacher_two.phone.label("invigilator_2_phone_number"),
            )
            .select_from(ExaminationSchedules)
            .join(Classes, Classes.id == ExaminationSchedules.class_id)
            .join(Subjects, Subjects.id == ExaminationSchedules.subject_id)
            .outerjoin(Rooms, Rooms.id == ExaminationSchedules.room_id)
            .outerjoin(teacher_one, teacher_one.id == ExaminationSchedules.invigilator_1_id)
            .outerjoin(teacher_two, teacher_two.id == ExaminationSchedules.invigilator_2_id)
            .where(ExaminationSchedules.id == examination_schedules_id)
        ).first()

        if not row:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Examination Schedules does not exist",
            )
        return ExaminationScheduleServices._build_response(row=row._asdict())

    @staticmethod
    def create(
        *,
        session: Session,
        examination_schedule: ExaminationScheduleCreate,
    ) -> ExaminationSchedulePublic:
        existing = session.exec(
            select(ExaminationSchedules).where(
                ExaminationSchedules.class_id == examination_schedule.class_id,
                ExaminationSchedules.subject_id == examination_schedule.subject_id,
            )
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Examination Schedules already exists.",
            )
        new_examination_schedules = ExaminationSchedules(**examination_schedule.dict())
        session.add(new_examination_schedules)
        session.commit()
        session.refresh(new_examination_schedules)

        return new_examination_schedules

    @staticmethod
    def update(
        *,
        session: Session,
        examination_schedule_id: uuid.UUID,
        examination_schedule_data: ExaminationScheduleUpdate,
    ) -> ExaminationSchedulePublic:
        examination_schedule = session.get(
            ExaminationSchedules, examination_schedule_id
        )
        if not examination_schedule:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Examination Schedule not found",
            )

        update_data = examination_schedule_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(examination_schedule, field, value)

        session.commit()
        return ExaminationSchedulePublic.model_validate(examination_schedule)

    @staticmethod
    def delete(
        *,
        session: Session,
        examination_schedule_id: uuid.UUID,
    ) -> ExaminationScheduleDeleteResponse:
        examination_schedule = session.get(
            ExaminationSchedules, examination_schedule_id
        )
        if not examination_schedule:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Examination Schedule not found",
            )

        if examination_schedule.status != StatusEnum.INACTIVE:
            examination_schedule.status = StatusEnum.INACTIVE
            session.commit()
            return ExaminationScheduleDeleteResponse(
                id=str(examination_schedule.id),
                message="Examination Schedule set to inactive",
            )

        session.delete(examination_schedule)
        session.commit()
        return ExaminationScheduleDeleteResponse(
            id=str(examination_schedule.id),
            message="Examination Schedule deleted successfully",
        )
